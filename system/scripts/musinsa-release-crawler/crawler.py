"""
Musinsa Release Crawler for Wacky Willy Brand Intelligence
===========================================================
무신사 발매판 API를 통해 브랜드별 발매 정보를 수집하고 엑셀로 저장.
발매 예정/진행 중/종료 상품의 브랜드, 상품명, 발매일시, D-Day, 가격 데이터를 포함.

Selenium 없이 API 직접 호출 방식.

사용법:
  python crawler.py                             # 탭 선택 대화형 모드
  python crawler.py --tab now --gender A        # NOW 탭 전체
  python crawler.py --tab upcoming --sort latest
  python crawler.py --list-tabs                 # 탭 목록 출력
"""

from __future__ import annotations

import os
import sys
import json
import argparse
import requests
from datetime import datetime, timezone

from io import BytesIO

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XlImage
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from PIL import Image as PILImage
    HAS_PILLOW = True
except ImportError:
    HAS_PILLOW = False


# ──────────────────────────────────────────────
# 1. 탭 & 필터 매핑
# ──────────────────────────────────────────────

# 탭 → sectionId 매핑 (무신사 발매판 고정 ID)
TABS = {
    "now":            {"id": 91,  "name": "NOW (현재 발매)"},
    "upcoming":       {"id": 92,  "name": "예정"},
    "popular-re":     {"id": 93,  "name": "인기 재발매"},
    "exclusive":      {"id": 94,  "name": "무신사 단독"},
    "sneakers":       {"id": 95,  "name": "스니커즈 캘린더"},
}

GENDERS = {
    "A": "전체",
    "M": "남성",
    "F": "여성",
}

SORTS = {
    "latest":  "최신순",
    "popular": "인기순",
}

# ──────────────────────────────────────────────
# 2. API 호출
# ──────────────────────────────────────────────

BASE_URL  = "https://api.musinsa.com/api2/hm/web/v1/pans/release/sections"
PAGE_URL  = "https://api.musinsa.com/api2/hm/web/v1/pans/release"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/131.0.0.0 Safari/537.36"
    ),
    "Accept": "application/json",
    "Referer": "https://www.musinsa.com/",
    "Origin": "https://www.musinsa.com",
}


def discover_tabs(gender: str = "A") -> dict:
    """발매 메인 페이지 API에서 탭 목록을 자동 탐색한다.
    성공하면 {tab_id: tab_name} 딕셔너리 반환, 실패하면 기본 TABS 반환.
    """
    try:
        resp = requests.get(
            PAGE_URL,
            params={"storeCode": "musinsa", "gf": gender},
            headers=HEADERS,
            timeout=20,
        )
        resp.raise_for_status()
        data = resp.json()
        modules = data.get("data", {}).get("modules", [])
        discovered = {}
        for mod in modules:
            if mod.get("type") == "TAB_OUTLINED":
                for item in mod.get("items", []):
                    tid = item.get("sectionId") or item.get("id")
                    tname = item.get("text", "") or item.get("label", "")
                    if tid and tname:
                        discovered[str(tid)] = tname
        if discovered:
            print(f"  [탭 탐색] {len(discovered)}개 탭 발견: {list(discovered.values())}")
            return discovered
    except Exception as e:
        print(f"  [탭 탐색] 자동 탐색 실패 ({e}), 기본값 사용")
    return {str(v["id"]): v["name"] for v in TABS.values()}


def fetch_release(section_id: int, gender: str, sort: str) -> list[dict]:
    """무신사 발매판 API에서 발매 상품 데이터를 가져온다."""
    url = f"{BASE_URL}/{section_id}"
    params = {
        "storeCode": "musinsa",
        "gf": gender,
        "sort": sort,
    }

    print(f"  API 호출: section={section_id}, gender={gender}, sort={sort}")

    try:
        resp = requests.get(url, params=params, headers=HEADERS, timeout=30)
        resp.raise_for_status()
        data = resp.json()
    except requests.RequestException as e:
        print(f"  [ERROR] API 호출 실패: {e}")
        return []

    items = []
    modules = data.get("data", {}).get("modules", [])

    for module in modules:
        mod_type = module.get("type", "")

        # 대형 발매 배너 (CAROUSEL_ONEROW_SNAPPING 안의 CONTENT_CAROUSEL_LARGE_RELEASE)
        if mod_type == "CAROUSEL_ONEROW_SNAPPING":
            for banner in module.get("items", []):
                if "RELEASE" in banner.get("type", ""):
                    item = parse_release_item(banner, item_type="banner")
                    if item:
                        items.append(item)

        # 상품 카루셀 (PRODUCT_CAROUSEL)
        elif mod_type in ("PRODUCT_CAROUSEL", "MULTICOLUMN"):
            for sub in module.get("items", []):
                if sub.get("type") == "PRODUCT_CAROUSEL":
                    item = parse_release_item(sub, item_type="product")
                    if item:
                        items.append(item)

    # CONTENT_CAROUSEL_LARGE_RELEASE가 최상위 모듈인 경우도 처리
    for module in modules:
        if "RELEASE" in module.get("type", "") and module.get("info"):
            item = parse_release_item(module, item_type="banner")
            if item:
                items.append(item)

    # 중복 제거 (동일 product_id)
    seen = set()
    unique = []
    for it in items:
        key = it.get("product_id") or it.get("title")
        if key and key not in seen:
            seen.add(key)
            unique.append(it)

    print(f"  수집 완료: {len(unique)}개 발매 상품")
    return unique


def parse_release_item(item: dict, item_type: str = "product") -> dict | None:
    """발매 아이템 데이터를 정규화한다."""
    try:
        info      = item.get("info", {})
        image     = item.get("image", {})
        on_click  = item.get("onClick", {})

        # 발매 타임스탬프 → 날짜 변환
        ts_ms = info.get("releaseTargetDate", 0)
        release_dt_str = ""
        release_date_str = ""
        dday = ""
        release_status = "확인필요"

        if ts_ms:
            release_dt = datetime.fromtimestamp(ts_ms / 1000, tz=timezone.utc).astimezone()
            now = datetime.now(tz=timezone.utc).astimezone()
            delta = (release_dt.date() - now.date()).days
            release_dt_str  = release_dt.strftime("%Y-%m-%d %H:%M")
            release_date_str = release_dt.strftime("%Y-%m-%d")

            if delta > 0:
                dday = f"D-{delta}"
                release_status = "예정"
            elif delta == 0:
                dday = "D-Day"
                release_status = "오늘 발매"
            else:
                dday = f"D+{abs(delta)}"
                release_status = "발매 완료"

        # 가격
        final_price   = info.get("finalPrice", 0)
        discount_ratio = info.get("discountRatio", 0)
        original_price = info.get("originalPrice", 0)
        if not original_price and discount_ratio and final_price:
            original_price = int(final_price / (1 - discount_ratio / 100))

        # 상품 URL
        product_url = ""
        if isinstance(on_click, dict):
            product_url = on_click.get("url", "")

        # 이미지 URL
        image_url = image.get("url", "") if isinstance(image, dict) else ""
        if image_url and image_url.startswith("//"):
            image_url = "https:" + image_url

        # 브랜드명 / 상품명 — API가 plain string 또는 {"text": "..."} 모두 반환
        def _text(val) -> str:
            if isinstance(val, dict):
                return val.get("text", "") or val.get("value", "")
            return str(val) if val else ""

        brand_name   = _text(info.get("brandName") or info.get("brand", ""))
        product_name = _text(
            info.get("productName") or info.get("title") or item.get("title", "")
        )
        sub_title    = info.get("subTitle", "")  # "오늘 오전 10:00" 형식 텍스트

        # 이미지 서브타이틀에서 발매 시간 텍스트 보완
        if sub_title and not release_dt_str:
            release_dt_str = sub_title

        return {
            "item_type":      item_type,          # "banner" | "product"
            "brand_name":     brand_name,
            "product_name":   product_name,
            "product_id":     str(item.get("id", "")),
            "release_datetime": release_dt_str,
            "release_date":   release_date_str,
            "dday":           dday,
            "release_status": release_status,
            "sub_title":      sub_title,
            "original_price": original_price,
            "final_price":    final_price,
            "discount_ratio": discount_ratio,
            "is_sold_out":    info.get("isSoldOut", False),
            "image_url":      image_url,
            "product_url":    product_url,
        }
    except Exception as e:
        print(f"  [WARN] 발매 아이템 파싱 실패: {e}")
        return None


# ──────────────────────────────────────────────
# 3. 엑셀 저장
# ──────────────────────────────────────────────

IMG_WIDTH_PX  = 70
IMG_HEIGHT_PX = 90
ROW_HEIGHT_PT = 70

STATUS_COLORS = {
    "예정":     "E8F4FD",   # 연파랑
    "오늘 발매": "FFF3CD",   # 연노랑
    "발매 완료": "F8F9FA",   # 연회색
    "확인필요":  "FFFFFF",
}

DDAY_FONT_COLORS = {
    "예정":     "0066CC",
    "오늘 발매": "CC6600",
    "발매 완료": "888888",
    "확인필요":  "333333",
}


def _download_thumbnail(url: str) -> bytes | None:
    """상품 썸네일을 다운로드한다."""
    if not url:
        return None
    try:
        if url.startswith("//"):
            url = "https:" + url
        resp = requests.get(url, headers=HEADERS, timeout=10)
        if resp.status_code == 200 and len(resp.content) > 1000:
            return resp.content
    except Exception:
        pass
    return None


def save_to_excel(items: list[dict], output_path: str, meta: dict, embed_images: bool = True):
    """발매 상품 데이터를 엑셀 파일로 저장한다."""
    if not HAS_OPENPYXL:
        print("  [WARN] openpyxl 미설치 — JSON으로 대체 저장합니다.")
        json_path = output_path.replace(".xlsx", ".json")
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump({"meta": meta, "items": items}, f, ensure_ascii=False, indent=2)
        print(f"  저장: {json_path}")
        return json_path

    wb = Workbook()
    ws = wb.active
    ws.title = "무신사 발매"

    can_embed = embed_images and HAS_OPENPYXL and HAS_PILLOW

    # 컬럼 정의
    columns = [
        ("이미지",      10),   # A
        ("브랜드",      18),   # B
        ("상품명",      38),   # C
        ("발매일시",    18),   # D
        ("D-Day",       8),    # E
        ("발매상태",    10),   # F
        ("원가",        12),   # G
        ("할인율",       8),   # H
        ("판매가",      12),   # I
        ("품절",         6),   # J
        ("상품 URL",    50),   # K
        ("이미지 URL",  55),   # L
    ]
    IMG_COL_IDX = 1   # A열

    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    last_col_letter = get_column_letter(len(columns))

    # 메타 헤더
    meta_text = (
        f"무신사 발매판 | 탭: {meta.get('tab_name', '')} | "
        f"성별: {meta.get('gender_name', '')} | 정렬: {meta.get('sort_name', '')} | "
        f"수집: {meta.get('crawled_at', '')}"
    )
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = meta_text
    ws["A1"].font = Font(bold=True, size=11)
    ws["A1"].fill = PatternFill("solid", fgColor="1A1A2E")
    ws["A1"].font = Font(bold=True, size=11, color="FFFFFF")

    # 컬럼 헤더
    header_font  = Font(bold=True, color="FFFFFF", size=10)
    header_fill  = PatternFill("solid", fgColor="16213E")
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx, (col_name, col_width) in enumerate(columns, 1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.font   = header_font
        cell.fill   = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # 이미지 다운로드
    image_data_map = {}
    if can_embed:
        print("  이미지 다운로드 중...")
        downloaded = 0
        for i, it in enumerate(items):
            img_bytes = _download_thumbnail(it.get("image_url", ""))
            if img_bytes:
                image_data_map[i] = img_bytes
                downloaded += 1
            if (i + 1) % 20 == 0:
                print(f"    {i + 1}/{len(items)} ({downloaded}장)")
        print(f"  이미지 다운로드 완료: {downloaded}/{len(items)}장")

    # 데이터 행
    for row_idx, it in enumerate(items, 3):
        data_idx   = row_idx - 3
        status     = it.get("release_status", "확인필요")
        row_bg     = STATUS_COLORS.get(status, "FFFFFF")
        dday_color = DDAY_FONT_COLORS.get(status, "333333")

        row_data = [
            "",                                                           # A: 이미지
            it.get("brand_name", ""),                                     # B: 브랜드
            it.get("product_name", ""),                                   # C: 상품명
            it.get("release_datetime", it.get("sub_title", "")),          # D: 발매일시
            it.get("dday", ""),                                           # E: D-Day
            status,                                                       # F: 발매상태
            it.get("original_price", ""),                                 # G: 원가
            f'{it["discount_ratio"]}%' if it.get("discount_ratio") else "",  # H: 할인율
            it.get("final_price", ""),                                    # I: 판매가
            "품절" if it.get("is_sold_out") else "",                      # J: 품절
            it.get("product_url", ""),                                    # K: 상품 URL
            it.get("image_url", ""),                                      # L: 이미지 URL
        ]

        WRAP_COLS = {2, 3, 11, 12}

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.fill   = PatternFill("solid", fgColor=row_bg)

            if col_idx in WRAP_COLS:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            elif col_idx == IMG_COL_IDX:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # D-Day 컬럼 강조
            if col_idx == 5:
                cell.font = Font(bold=True, color=dday_color, size=10)
            # 가격 포맷
            elif col_idx in (7, 9):
                cell.number_format = "#,##0"
            # URL 링크 스타일
            elif col_idx in (11, 12):
                cell.font = Font(color="0066CC", underline="single", size=9)
            # 품절 빨간 표시
            elif col_idx == 10 and it.get("is_sold_out"):
                cell.font = Font(color="CC0000", bold=True)

        if can_embed:
            ws.row_dimensions[row_idx].height = ROW_HEIGHT_PT

        # 이미지 삽입
        if can_embed and data_idx in image_data_map:
            try:
                img_bytes = image_data_map[data_idx]
                pil_img   = PILImage.open(BytesIO(img_bytes))
                pil_img.thumbnail((IMG_WIDTH_PX, IMG_HEIGHT_PX), PILImage.LANCZOS)
                resized   = BytesIO()
                pil_img.save(resized, format="JPEG", quality=85)
                resized.seek(0)
                xl_img = XlImage(resized)
                ws.add_image(xl_img, f"A{row_idx}")
            except Exception:
                pass

    # 자동 필터 & 틀 고정
    ws.auto_filter.ref = f"A2:{last_col_letter}{len(items) + 2}"
    ws.freeze_panes   = "B3"

    wb.save(output_path)
    print(f"  저장: {output_path}")
    return output_path


def save_to_json(items: list[dict], output_path: str, meta: dict):
    payload = {
        "meta": meta,
        "items": items,
        "count": len(items),
    }
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    print(f"  저장: {output_path}")
    return output_path


# ──────────────────────────────────────────────
# 4. CLI
# ──────────────────────────────────────────────

def list_tabs():
    print("\n[발매 탭 목록]")
    for key, val in TABS.items():
        print(f"  {key:<14} (sectionId={val['id']})  →  {val['name']}")
    print("\n[성별]")
    for k, v in GENDERS.items():
        print(f"  {k}: {v}")
    print("\n[정렬]")
    for k, v in SORTS.items():
        print(f"  {k}: {v}")


def main():
    parser = argparse.ArgumentParser(description="무신사 발매판 크롤러")
    parser.add_argument("--tab",        default="now",
                        choices=list(TABS.keys()), help="발매 탭")
    parser.add_argument("--section-id", default=None, type=int,
                        help="sectionId 직접 지정 (--tab 보다 우선)")
    parser.add_argument("--gender",     default="A",
                        choices=GENDERS.keys(), help="성별 필터")
    parser.add_argument("--sort",       default="latest",
                        choices=SORTS.keys(), help="정렬 (latest/popular)")
    parser.add_argument("--output-dir", default=None, help="저장 디렉토리")
    parser.add_argument("--format",     default="xlsx",
                        choices=["xlsx", "json", "both"], help="저장 형식")
    parser.add_argument("--no-images",  action="store_true",
                        help="엑셀 이미지 삽입 비활성화 (빠른 저장)")
    parser.add_argument("--list-tabs",  action="store_true",
                        help="탭 목록 출력")
    args = parser.parse_args()

    if args.list_tabs:
        list_tabs()
        return

    # sectionId 결정
    if args.section_id:
        section_id = args.section_id
        tab_name   = f"sectionId={section_id}"
    else:
        section_id = TABS[args.tab]["id"]
        tab_name   = TABS[args.tab]["name"]

    # 저장 디렉토리
    if args.output_dir:
        output_dir = args.output_dir
    else:
        script_dir   = os.path.dirname(os.path.abspath(__file__))
        project_root = os.path.dirname(os.path.dirname(os.path.dirname(script_dir)))
        today        = datetime.now().strftime("%Y%m%d")
        output_dir   = os.path.join(
            project_root, "workspace", "musinsa-release", f"release_{today}"
        )

    os.makedirs(output_dir, exist_ok=True)

    print(f"\n{'='*60}")
    print(f"  무신사 발매판 크롤러")
    print(f"  탭: {tab_name} | 성별: {GENDERS[args.gender]} | 정렬: {SORTS[args.sort]}")
    print(f"{'='*60}\n")

    items = fetch_release(section_id, args.gender, args.sort)

    if not items:
        print("\n수집된 발매 상품이 없습니다. sectionId를 확인하거나 --list-tabs로 탭 목록을 확인하세요.")
        return

    meta = {
        "tab":        args.tab,
        "tab_name":   tab_name,
        "section_id": section_id,
        "gender":     args.gender,
        "gender_name": GENDERS[args.gender],
        "sort":       args.sort,
        "sort_name":  SORTS[args.sort],
        "total_items": len(items),
        "crawled_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
    }

    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    base_name = f"musinsa_release_{args.tab}_{args.gender}_{timestamp}"
    embed     = not args.no_images

    if args.format in ("xlsx", "both"):
        xlsx_path = os.path.join(output_dir, f"{base_name}.xlsx")
        save_to_excel(items, xlsx_path, meta, embed_images=embed)

    if args.format in ("json", "both"):
        json_path = os.path.join(output_dir, f"{base_name}.json")
        save_to_json(items, json_path, meta)

    # 결과 요약
    result = {
        "status":      "success",
        "total_items": len(items),
        "output_dir":  output_dir,
        "files":       [],
        "meta":        meta,
    }
    if args.format in ("xlsx", "both"):
        result["files"].append(xlsx_path)
    if args.format in ("json", "both"):
        result["files"].append(json_path)

    print(f"\n{'='*60}")
    print(f"  수집 완료: 총 {len(items)}개 발매 상품")
    upcoming = [it for it in items if it.get("release_status") == "예정"]
    today_r  = [it for it in items if it.get("release_status") == "오늘 발매"]
    done     = [it for it in items if it.get("release_status") == "발매 완료"]
    print(f"  예정: {len(upcoming)}개 | 오늘 발매: {len(today_r)}개 | 발매 완료: {len(done)}개")
    print(f"  저장 위치: {output_dir}")
    print(f"{'='*60}")

    result_path = os.path.join(output_dir, "_result.json")
    with open(result_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)


if __name__ == "__main__":
    main()
