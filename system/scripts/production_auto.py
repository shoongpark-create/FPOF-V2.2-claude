#!/usr/bin/env python3
"""
Wacky Willy 26SS 생산관리 자동화 — VBA→Python 포팅
====================================================
VBA module_production-automation.bas + module_auto-refresh.bas 로직을
openpyxl 기반 Python 스크립트로 1:1 재구현.

사용법:
  python production_auto.py                     # 주간전체갱신 (기본)
  python production_auto.py --erp               # ERP 데이터만 갱신
  python production_auto.py --supplier          # 협력사 데이터만 반영
  python production_auto.py --report            # 보고서만 갱신
  python production_auto.py --file ERP파일.xlsx  # ERP 수동 지정
"""

import argparse
import glob
import os
import shutil
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, numbers
from openpyxl.utils import get_column_letter

# ── 경로 ────────────────────────────────────────────────────
BASE_DIR = Path(__file__).resolve().parent.parent.parent / "workspace" / "26SS" / "sourcing"
MASTER_FILE = BASE_DIR / "W_26SS_생산관리현황.xlsx"
SUPPLIER_DIR = BASE_DIR / "협력사입력"
ERP_DIR = BASE_DIR / "ERP추출"
ARCHIVE_DIR = SUPPLIER_DIR / "archive"

# ── 시트명 ──────────────────────────────────────────────────
WS_MASTER = "생산현황"
WS_ACC = "ACC"
WS_ISSUE = "이슈관리"
WS_DASHBOARD = "종합현황판"
WS_WEEKLY = "주간보고"
WS_MONTHLY = "월간보고"
WS_D_ORDER = "_발주"
WS_D_RECV = "_입고"
WS_D_SALES = "_판매"
WS_D_DEFECT = "_불량"

# ── 생산현황 컬럼 (1-based, VBA 상수 그대로) ────────────────
HDR_ROW = 7
DATA_START = 9

C_DIV = 1
C_SEASON = 5
C_ITEM = 6
C_STYLENO = 8
C_COLOR = 14
C_PLAN_QTY = 25
C_SUPPLIER = 28
C_CONFIRMED = 32

# CO — 협력사
C_CO_COUNTRY = 40
C_CO_REGION = 41
C_CO_FACTORY = 42

# 스케줄-원자재 — 협력사 (col 43~69)
C_SCH_START = 43   # 원단처
C_SCH_END = 69     # 진도율 (메인)

# 스케줄-QC/PP — 협력사 (col 70~89)
C_QCPP_START = 70  # 성적서 발행일
C_QCPP_END = 89    # 진도율 (PP)

# 생산진행 — 협력사 (col 90~98)
C_MAT = 90
C_MAT_CONFIRM = 91
C_MAT_RECV = 92
C_CUT = 93
C_OUTSRC_PRINT = 94
C_SEW = 95
C_OUTSRC_WASH = 96
C_FINISH = 97
C_PACK = 98

# 입고예정 — 협력사 (col 99~112)
C_EXP_SD = 99
C_EXP_1D = 100
C_EXP_1Q = 101
C_EXP_TOT = 112

# ── 협력사 양식 → 생산현황 컬럼 매핑 (양식 col → 생산현황 col) ──
# 양식 col 1=STYLE NO, 2=COLOR (키), 3~ 데이터
SUPPLIER_COL_MAP = {
    # CO (양식 col 3~5 → 생산현황 col 40~42)
    3: 40, 4: 41, 5: 42,
    # 스케줄-원자재 (양식 col 6~32 → 생산현황 col 43~69)
    **{i: 43 + (i - 6) for i in range(6, 33)},
    # 스케줄-QC/PP (양식 col 33~52 → 생산현황 col 70~89)
    **{i: 70 + (i - 33) for i in range(33, 53)},
    # 생산진행 (양식 col 53~61 → 생산현황 col 90~98)
    **{i: 90 + (i - 53) for i in range(53, 62)},
    # 입고예정 (양식 col 62~75 → 생산현황 col 99~112)
    **{i: 99 + (i - 62) for i in range(62, 76)},
}

# 실입고 — ERP 자동
C_ACT_DATE = 114
C_ACT_QTY = 115
C_ACT_TOT = 126
C_ACT_RATE = 127
C_FIRST_RECV = 132
C_ONTIME = 133
C_LEAD = 134
C_SALES = 136
C_DEFECT = 137
C_DEF_RATE = 138

# ERP 입고데이터 컬럼 (원본 기준)
ERP_R_DATE = 1
ERP_R_ITEM = 9
ERP_R_COLOR = 13
ERP_R_QTY = 29

# ERP 발주데이터 컬럼 (원본 기준)
ERP_O_SEASON = 4
ERP_O_ITEM = 7
ERP_O_COLOR = 12
ERP_O_QTY = 50


# ══════════════════════════════════════════════════════════════
#  유틸
# ══════════════════════════════════════════════════════════════

def cell_val(ws, row, col):
    """셀 값 안전하게 읽기."""
    v = ws.cell(row=row, column=col).value
    return v if v is not None else ""


def cell_str(ws, row, col):
    return str(cell_val(ws, row, col)).strip()


def cell_num(ws, row, col):
    v = cell_val(ws, row, col)
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0


def cell_date(ws, row, col):
    """날짜 값 반환 (date 또는 None)."""
    v = cell_val(ws, row, col)
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def last_row(ws, col):
    """해당 컬럼의 마지막 데이터 행."""
    for r in range(ws.max_row, 0, -1):
        if ws.cell(row=r, column=col).value is not None:
            return r
    return 0


def find_col(ws, hdr_row, txt, exact=False):
    """헤더 행에서 컬럼 찾기."""
    for c in range(1, ws.max_column + 1):
        v = str(ws.cell(row=hdr_row, column=c).value or "").strip()
        if exact:
            if v.upper() == txt.upper():
                return c
        else:
            if txt.upper() in v.upper():
                return c
    return 0


def ensure_sheet(wb, name, hidden=False):
    """시트 존재 확인, 없으면 생성."""
    if name in wb.sheetnames:
        return wb[name]
    ws = wb.create_sheet(title=name)
    if hidden:
        ws.sheet_state = "veryHidden"
    return ws


def has_sheet(wb, name):
    return name in wb.sheetnames


def build_index(ws):
    """생산현황 시트에서 STYLE NO|COLOR → 행번호 인덱스."""
    idx = {}
    lr = last_row(ws, C_STYLENO)
    for r in range(DATA_START, lr + 1):
        sn = cell_str(ws, r, C_STYLENO)
        cc = cell_str(ws, r, C_COLOR)
        if sn and cc:
            k = f"{sn}|{cc}".upper()
            if k not in idx:
                idx[k] = r
    return idx


# ══════════════════════════════════════════════════════════════
#  Step 1: ERP 데이터 갱신
# ══════════════════════════════════════════════════════════════

def paste_erp_direct(src_ws, wb, tgt_name, title, item_col, color_col):
    """ERP 원본 시트 → 숨김 데이터 시트로 복사 + SKU2 생성."""
    ws = ensure_sheet(wb, tgt_name, hidden=True)

    # 원본을 메모리로 일괄 읽기 (read_only 시트에서 셀 단위보다 훨씬 빠름)
    rows = list(src_ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return 0

    src_lr = min(len(rows), 50000)
    src_lc = max(len(r) for r in rows[:src_lr])

    # 대상 클리어 — 시트 삭제 후 재생성이 셀 단위 클리어보다 빠름
    idx = wb.sheetnames.index(tgt_name)
    del wb[tgt_name]
    ws = wb.create_sheet(title=tgt_name, index=idx)
    ws.sheet_state = "veryHidden"

    # Row 1: 타이틀
    ws.cell(row=1, column=1, value=f"{title} ({src_lr - 1}행)")
    ws.cell(row=1, column=2, value=f"UPDATED: {datetime.now():%Y-%m-%d %H:%M}")

    # Row 3: SKU2 헤더 + 원본 헤더 (1열 오프셋)
    hdr = rows[0]
    ws.cell(row=3, column=1, value="SKU2")
    for c in range(len(hdr)):
        ws.cell(row=3, column=c + 2, value=hdr[c])

    # Row 4+: 데이터 일괄 쓰기 + SKU2 동시 생성
    t_item_idx = item_col - 1   # 0-based index into row tuple
    t_color_idx = color_col - 1
    for i in range(1, src_lr):
        row_data = rows[i]
        tgt_row = i + 3  # row 4, 5, 6, ...

        # SKU2
        itm = str(row_data[t_item_idx] or "").strip() if t_item_idx < len(row_data) else ""
        clr = str(row_data[t_color_idx] or "").strip() if t_color_idx < len(row_data) else ""
        if itm:
            ws.cell(row=tgt_row, column=1, value=(itm + clr).upper())

        # 원본 데이터 (1열 오프셋)
        for c in range(len(row_data)):
            ws.cell(row=tgt_row, column=c + 2, value=row_data[c])

    return src_lr - 1


def erp_갱신_자동(wb):
    """ERP추출/ 폴더에서 최신 erp_data_*.xlsx 파일을 찾아 갱신."""
    files = sorted(ERP_DIR.glob("erp_data_*.xlsx"), key=os.path.getmtime, reverse=True)
    if not files:
        print(f"  ⚠ ERP추출 폴더에 erp_data_*.xlsx 없음: {ERP_DIR}")
        return

    latest = files[0]
    print(f"  ERP 파일: {latest.name}")

    wb_erp = openpyxl.load_workbook(latest, read_only=True, data_only=True)
    msg = []

    if "발주데이터" in wb_erp.sheetnames:
        n = paste_erp_direct(wb_erp["발주데이터"], wb, WS_D_ORDER,
                             "26SS 발주 데이터", ERP_O_ITEM, ERP_O_COLOR)
        msg.append(f"  발주: {n}행")

    if "입고데이터" in wb_erp.sheetnames:
        n = paste_erp_direct(wb_erp["입고데이터"], wb, WS_D_RECV,
                             "입고 데이터", ERP_R_ITEM, ERP_R_COLOR)
        msg.append(f"  입고: {n}행")

    wb_erp.close()
    for m in msg:
        print(m)


def erp_갱신_수동(wb, filepath):
    """지정된 ERP 파일로 갱신."""
    fp = Path(filepath)
    if not fp.exists():
        print(f"  ⚠ 파일 없음: {fp}")
        return

    print(f"  ERP 파일 (수동): {fp.name}")
    wb_erp = openpyxl.load_workbook(fp, read_only=True, data_only=True)

    if "발주데이터" in wb_erp.sheetnames:
        paste_erp_direct(wb_erp["발주데이터"], wb, WS_D_ORDER,
                         "26SS 발주 데이터", ERP_O_ITEM, ERP_O_COLOR)

    if "입고데이터" in wb_erp.sheetnames:
        paste_erp_direct(wb_erp["입고데이터"], wb, WS_D_RECV,
                         "입고 데이터", ERP_R_ITEM, ERP_R_COLOR)

    wb_erp.close()
    print("  ERP 수동 갱신 완료")


# ══════════════════════════════════════════════════════════════
#  Step 2: 협력사 데이터 반영
# ══════════════════════════════════════════════════════════════

def apply_supplier_data(src_ws, tgt_ws, idx):
    """협력사 종합 입력 시트 → 생산현황 반영.

    양식 컬럼 매핑(SUPPLIER_COL_MAP)을 사용하여
    CO, 스케줄(원자재/QC·PP), 생산진행, 입고예정 전체를 한 번에 반영.
    """
    cnt = 0
    lr = src_ws.max_row

    for r in range(5, lr + 1):  # row 5부터 데이터 (row 1~4 헤더)
        sn = str(src_ws.cell(row=r, column=1).value or "").strip()
        cc = str(src_ws.cell(row=r, column=2).value or "").strip()
        if not sn or not cc:
            continue

        k = f"{sn}|{cc}".upper()
        if k not in idx:
            continue

        tr = idx[k]
        changed = False

        for src_col, tgt_col in SUPPLIER_COL_MAP.items():
            v = src_ws.cell(row=r, column=src_col).value
            if v is not None and v != "":
                tgt_ws.cell(row=tr, column=tgt_col, value=v)
                changed = True

        # 입고예정 수량 합계 재계산 (col 101, 103, 105, 107, 109, 111)
        if changed:
            tot = sum(cell_num(tgt_ws, tr, C_EXP_1Q + (s - 1) * 2) for s in range(1, 7))
            if tot > 0:
                tgt_ws.cell(row=tr, column=C_EXP_TOT, value=tot)
            cnt += 1

    return cnt


# ── 구버전 양식 호환 (생산진행/입고예정 2시트 방식) ──

def apply_progress_legacy(src_ws, tgt_ws, idx):
    """구버전 협력사 '생산진행' 시트 → 생산현황 반영."""
    cnt = 0
    col_map = {3: C_MAT, 4: C_CUT, 5: C_SEW, 6: C_FINISH, 7: C_PACK}

    for r in range(2, src_ws.max_row + 1):
        sn = str(src_ws.cell(row=r, column=1).value or "").strip()
        cc = str(src_ws.cell(row=r, column=2).value or "").strip()
        k = f"{sn}|{cc}".upper()

        if k in idx:
            tr = idx[k]
            for sc, tc in col_map.items():
                v = src_ws.cell(row=r, column=sc).value
                if v is not None and v != "":
                    tgt_ws.cell(row=tr, column=tc, value=v)
            cnt += 1
    return cnt


def apply_delivery_legacy(src_ws, tgt_ws, idx):
    """구버전 협력사 '입고예정' 시트 → 생산현황 반영."""
    cnt = 0

    for r in range(2, src_ws.max_row + 1):
        sn = str(src_ws.cell(row=r, column=1).value or "").strip()
        cc = str(src_ws.cell(row=r, column=2).value or "").strip()
        k = f"{sn}|{cc}".upper()

        if k in idx:
            tr = idx[k]
            ship_no = int(cell_num(src_ws, r, 3))
            if 1 <= ship_no <= 6:
                dc = C_EXP_1D + (ship_no - 1) * 2
                qc = C_EXP_1Q + (ship_no - 1) * 2
                v_date = src_ws.cell(row=r, column=4).value
                v_qty = src_ws.cell(row=r, column=5).value
                if v_date is not None and v_date != "":
                    tgt_ws.cell(row=tr, column=dc, value=v_date)
                if v_qty is not None and v_qty != "":
                    tgt_ws.cell(row=tr, column=qc, value=v_qty)
                tot = sum(cell_num(tgt_ws, tr, C_EXP_1Q + (s - 1) * 2) for s in range(1, 7))
                tgt_ws.cell(row=tr, column=C_EXP_TOT, value=tot)
                cnt += 1
    return cnt


def 협력사데이터반영(wb):
    """협력사입력/ 폴더 스캔 → 생산현황 반영 → archive 이동.

    새 양식('협력사입력' 단일 시트)과 구 양식('생산진행'+'입고예정' 2시트) 모두 호환.
    """
    if not SUPPLIER_DIR.exists():
        print(f"  ⚠ 폴더 없음: {SUPPLIER_DIR}")
        return

    if not has_sheet(wb, WS_MASTER):
        print(f"  ⚠ '{WS_MASTER}' 시트 없음")
        return

    ws = wb[WS_MASTER]
    idx = build_index(ws)

    files = sorted(SUPPLIER_DIR.glob("협력사입력_*.xlsx"))
    # 템플릿 제외
    files = [f for f in files if "템플릿" not in f.name and "~$" not in f.name]

    if not files:
        print("  협력사 파일 없음 (스킵)")
        return

    ARCHIVE_DIR.mkdir(exist_ok=True)
    total = 0

    for fp in files:
        wb_s = openpyxl.load_workbook(fp, read_only=True, data_only=True)
        n = 0

        if "협력사입력" in wb_s.sheetnames:
            # 새 양식: 단일 시트에 CO/스케줄/생산진행/입고예정 통합
            n += apply_supplier_data(wb_s["협력사입력"], ws, idx)
            fmt = "신규"
        else:
            # 구 양식: 생산진행 + 입고예정 2시트
            if "생산진행" in wb_s.sheetnames:
                n += apply_progress_legacy(wb_s["생산진행"], ws, idx)
            if "입고예정" in wb_s.sheetnames:
                n += apply_delivery_legacy(wb_s["입고예정"], ws, idx)
            fmt = "구버전"

        wb_s.close()

        # archive 이동
        dest = ARCHIVE_DIR / fp.name
        try:
            shutil.move(str(fp), str(dest))
        except Exception:
            pass

        total += n
        print(f"  {fp.name}: {n}건 반영 ({fmt} 양식)")

    print(f"  협력사 합계: 파일 {len(files)}개 / {total}건")


# ══════════════════════════════════════════════════════════════
#  Step 3-a: 실입고 자동 반영 (_입고 → 생산현황)
# ══════════════════════════════════════════════════════════════

def 실입고자동반영(wb):
    """_입고 시트 데이터 → 생산현황 실입고 컬럼."""
    if not has_sheet(wb, WS_D_RECV):
        return

    ws_m = wb[WS_MASTER]
    ws_r = wb[WS_D_RECV]

    # 컬럼 위치 탐색
    c_key = 0
    c_qty = 0
    c_date = 0
    for hr in range(1, 4):
        if not c_key:
            c_key = find_col(ws_r, hr, "SKU2", exact=True)
        if not c_qty:
            c_qty = find_col(ws_r, hr, "계")
        if not c_date:
            c_date = find_col(ws_r, hr, "입고일자")

    if not c_key:
        c_key = 1
    if not c_qty:
        c_qty = 30
    if not c_date:
        c_date = 2

    # _입고 집계: SKU2 → (수량합계, 최초입고일)
    recv_dict = {}  # key → [qty_sum, first_date]
    lr = last_row(ws_r, c_key)
    data_start = 4

    if lr < data_start:
        return

    for r in range(data_start, lr + 1):
        k = cell_str(ws_r, r, c_key).upper()
        if not k or k == "0":
            continue

        q = cell_num(ws_r, r, c_qty)
        d = cell_date(ws_r, r, c_date)

        if k in recv_dict:
            recv_dict[k][0] += q
            if d and (recv_dict[k][1] is None or d < recv_dict[k][1]):
                recv_dict[k][1] = d
        else:
            recv_dict[k] = [q, d]

    # 생산현황에 반영
    lr_m = last_row(ws_m, C_STYLENO)
    cnt = 0

    for r in range(DATA_START, lr_m + 1):
        sn = cell_str(ws_m, r, C_STYLENO)
        cc = cell_str(ws_m, r, C_COLOR)
        if not sn or not cc:
            continue

        k = (sn + cc).upper()
        if k in recv_dict:
            qty, first_d = recv_dict[k]
            ws_m.cell(row=r, column=C_ACT_QTY, value=qty)
            ws_m.cell(row=r, column=C_ACT_TOT, value=qty)

            if first_d:
                ws_m.cell(row=r, column=C_ACT_DATE, value=first_d)
                ws_m.cell(row=r, column=C_FIRST_RECV, value=first_d)

            pq = cell_num(ws_m, r, C_PLAN_QTY)
            if pq > 0:
                rate = qty / pq
                c = ws_m.cell(row=r, column=C_ACT_RATE, value=rate)
                c.number_format = "0%"

            cnt += 1

    print(f"  실입고 반영: {cnt}건")


# ══════════════════════════════════════════════════════════════
#  Step 3-b: 불량 자동 반영 (_판매 + _불량 → 생산현황)
# ══════════════════════════════════════════════════════════════

def build_sum_dict(ws, key_hdr, qty_hdr):
    """시트에서 키-수량 합산 딕셔너리 생성."""
    d = {}
    ck = cq = 0
    found_row = 2

    for hr in range(1, 5):
        if not ck:
            ck = find_col(ws, hr, key_hdr)
        if not cq:
            cq = find_col(ws, hr, qty_hdr)
        if ck and cq:
            found_row = hr
            break

    if not ck or not cq:
        return d

    start_r = max(found_row, 2) + 1
    lr = last_row(ws, ck)

    for r in range(start_r, lr + 1):
        k = cell_str(ws, r, ck).upper()
        if k:
            v = cell_num(ws, r, cq)
            d[k] = d.get(k, 0) + v

    return d


def 불량자동반영(wb):
    """_판매 + _불량 → 생산현황 판매/불량/불량률."""
    ws_m = wb[WS_MASTER]

    s_dict = {}
    if has_sheet(wb, WS_D_SALES):
        s_dict = build_sum_dict(wb[WS_D_SALES], "품번", "수량")

    d_dict = {}
    if has_sheet(wb, WS_D_DEFECT):
        d_dict = build_sum_dict(wb[WS_D_DEFECT], "품번", "수량")

    lr = last_row(ws_m, C_STYLENO)
    cnt = 0

    for r in range(DATA_START, lr + 1):
        sn = cell_str(ws_m, r, C_STYLENO)
        cc = cell_str(ws_m, r, C_COLOR)
        if not sn or not cc:
            continue

        k = (sn + cc).upper()

        if k in s_dict:
            ws_m.cell(row=r, column=C_SALES, value=s_dict[k])
        if k in d_dict:
            ws_m.cell(row=r, column=C_DEFECT, value=d_dict[k])

        sv = cell_num(ws_m, r, C_SALES)
        dv = cell_num(ws_m, r, C_DEFECT)
        if sv > 0:
            c = ws_m.cell(row=r, column=C_DEF_RATE, value=dv / sv)
            c.number_format = "0.0%"
            cnt += 1

    print(f"  불량 반영: {cnt}건")


# ══════════════════════════════════════════════════════════════
#  Step 3-c: 적기 입고 판정
# ══════════════════════════════════════════════════════════════

def 적기입고판정(wb):
    """확정납기 vs 최초입고일 → 적기/지연/미입고 판정."""
    ws = wb[WS_MASTER]
    lr = last_row(ws, C_STYLENO)
    today = date.today()
    cnt = 0

    for r in range(DATA_START, lr + 1):
        if not cell_val(ws, r, C_STYLENO):
            continue

        conf_date = cell_date(ws, r, C_CONFIRMED)
        recv_date = cell_date(ws, r, C_FIRST_RECV)

        if conf_date and recv_date:
            diff = (recv_date - conf_date).days
            ws.cell(row=r, column=C_LEAD, value=diff)

            if diff <= 0:
                ws.cell(row=r, column=C_ONTIME, value="적기")
            elif diff <= 7:
                ws.cell(row=r, column=C_ONTIME, value="지연(7일내)")
            else:
                ws.cell(row=r, column=C_ONTIME, value="지연")
            cnt += 1

        elif conf_date and conf_date < today:
            ws.cell(row=r, column=C_ONTIME, value="미입고")
            ws.cell(row=r, column=C_LEAD, value=(today - conf_date).days)
            cnt += 1

    print(f"  적기 판정: {cnt}건")


# ══════════════════════════════════════════════════════════════
#  Step 4: 보고서 갱신 (종합현황판 + 주간보고 + 월간보고)
# ══════════════════════════════════════════════════════════════

BOLD = Font(bold=True)
TITLE_FONT = Font(bold=True, size=14)
SEC_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=9)
DARK_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
GREEN_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
RED_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
ALERT_FILL = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
ALERT_FONT = Font(bold=True, color="FFFFFF")
KPI_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
KPI_FONT = Font(bold=True, size=16)
KPI_LABEL = Font(size=9, color="666666")
GRAY_FONT = Font(color="999999", size=9)
NUM_FMT = "#,##0"
PCT_FMT = "0.0%"
PCT2_FMT = "0.00%"
DATE_FMT = "YYYY-MM-DD"
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

# 공정 컬럼 (생산현황 기준)
PROCESS_COLS = {
    "원자재입고": 90, "자재확정": 91, "자재입고": 92, "재단": 93,
    "외주(프린트)": 94, "봉제": 95, "외주(워싱)": 96, "완성": 97, "수납": 98,
}
SCHED_COLS = {
    "BT CF": 58, "메인 CF": 68, "QC AP": 80, "PP": 87,
}


def _clear_sheet(ws):
    # 병합 셀 해제
    for merge in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merge))
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for c in row:
            c.value = None
            c.font = Font()
            c.fill = PatternFill()
            c.alignment = Alignment()
            c.number_format = "General"


def _write_header_row(ws, row, col_start, headers, fill=HEADER_FILL, font=HEADER_FONT):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=col_start + i, value=h)
        c.font = font
        c.fill = fill
        c.alignment = CENTER


def _write_kpi_card(ws, row, col, label, value, fmt="", sublabel=""):
    ws.cell(row=row, column=col, value=label).font = KPI_LABEL
    c = ws.cell(row=row + 1, column=col, value=value)
    c.font = KPI_FONT
    c.fill = KPI_FILL
    c.alignment = CENTER
    if fmt:
        c.number_format = fmt
    if sublabel:
        ws.cell(row=row + 2, column=col, value=sublabel).font = GRAY_FONT


def _ontime_fill(status):
    if status == "적기":
        return GREEN_FILL
    elif status == "지연(7일내)":
        return YELLOW_FILL
    elif status == "지연":
        return RED_FILL
    elif status == "미입고":
        return PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    return PatternFill()


# ── 데이터 수집 (한 번 읽고 3개 시트에서 재사용) ──

def _collect_data(ws_m):
    """생산현황에서 전체 데이터를 수집하여 분석 구조체 반환."""
    lr = last_row(ws_m, C_STYLENO)
    today = date.today()

    records = []
    for r in range(DATA_START, lr + 1):
        sn = cell_str(ws_m, r, C_STYLENO)
        if not sn:
            continue

        rec = {
            "row": r,
            "div": cell_str(ws_m, r, C_DIV),
            "gubun": cell_str(ws_m, r, 4),
            "item": cell_str(ws_m, r, C_ITEM),
            "style": sn,
            "color": cell_str(ws_m, r, C_COLOR),
            "sourcer": cell_str(ws_m, r, 7),
            "supplier": cell_str(ws_m, r, C_SUPPLIER),
            "country": cell_str(ws_m, r, C_CO_COUNTRY),
            "plan_qty": cell_num(ws_m, r, C_PLAN_QTY),
            "recv_qty": cell_num(ws_m, r, C_ACT_TOT),
            "sales": cell_num(ws_m, r, C_SALES),
            "defect": cell_num(ws_m, r, C_DEFECT),
            "cost": cell_num(ws_m, r, 33),        # 사입가
            "price": cell_num(ws_m, r, 35),        # 판매가
            "mfg_amt": cell_num(ws_m, r, 34),      # 제조금액
            "supply_amt": cell_num(ws_m, r, 36),   # 공급금액
            "markup": cell_num(ws_m, r, 37),       # M/UP
            "cost_rate": cell_num(ws_m, r, 38),    # 원가율
            "conf_date": cell_date(ws_m, r, C_CONFIRMED),
            "recv_date": cell_date(ws_m, r, C_FIRST_RECV),
            "ontime": cell_str(ws_m, r, C_ONTIME),
            "lead": cell_num(ws_m, r, C_LEAD),
            "recv_rate": cell_num(ws_m, r, C_ACT_RATE),
            "exp_qty": cell_num(ws_m, r, C_EXP_TOT),
        }

        # 스케줄 진행 체크
        for name, col in SCHED_COLS.items():
            rec[f"sched_{name}"] = cell_val(ws_m, r, col) not in (None, "")

        # 생산공정 진행 체크
        for name, col in PROCESS_COLS.items():
            rec[f"proc_{name}"] = cell_val(ws_m, r, col) not in (None, "")

        # 월 분류 (확정납기 기준)
        if rec["conf_date"]:
            rec["conf_month"] = rec["conf_date"].strftime("%Y-%m")
        else:
            rec["conf_month"] = None

        records.append(rec)

    return records


def _group_by(records, key):
    """레코드를 키별로 그룹화."""
    groups = {}
    for r in records:
        k = r.get(key, "")
        if k not in groups:
            groups[k] = []
        groups[k].append(r)
    return groups


def _agg(records):
    """그룹 집계."""
    st = len(records)
    plan = sum(r["plan_qty"] for r in records)
    recv = sum(r["recv_qty"] for r in records)
    sales = sum(r["sales"] for r in records)
    defect = sum(r["defect"] for r in records)
    mfg = sum(r["mfg_amt"] for r in records)
    supply = sum(r["supply_amt"] for r in records)
    ontime = sum(1 for r in records if r["ontime"] == "적기")
    delay = sum(1 for r in records if r["ontime"] in ("지연", "지연(7일내)"))
    unreceived = sum(1 for r in records if r["ontime"] == "미입고")
    return {
        "st": st, "plan": plan, "recv": recv, "sales": sales, "defect": defect,
        "mfg_amt": mfg, "supply_amt": supply,
        "recv_rate": recv / plan if plan > 0 else 0,
        "def_rate": defect / sales if sales > 0 else 0,
        "ontime": ontime, "delay": delay, "unreceived": unreceived,
        "ontime_rate": ontime / (ontime + delay) if (ontime + delay) > 0 else 0,
    }


# ══════════════════════════════════════════════════════════════
#  종합현황판 — 실무자용 (문제 판단 + 시각화)
# ══════════════════════════════════════════════════════════════

def 종합현황판갱신(wb, records=None):
    ws_d = ensure_sheet(wb, WS_DASHBOARD)
    ws_m = wb[WS_MASTER]
    _clear_sheet(ws_d)

    if records is None:
        records = _collect_data(ws_m)

    total = _agg(records)
    today = date.today()

    # ── 타이틀 ──
    ws_d.merge_cells("B2:H2")
    ws_d.cell(row=2, column=2, value="◆ Wacky Willy 26SS 종합생산현황판").font = TITLE_FONT
    ws_d.cell(row=2, column=12, value=f"갱신: {datetime.now():%Y-%m-%d %H:%M}").font = GRAY_FONT

    # ── Section 1: 핵심 KPI 카드 ──
    _write_kpi_card(ws_d, 4, 2, "총 발주", total["st"], sublabel="ST")
    _write_kpi_card(ws_d, 4, 4, "기획 수량", total["plan"], fmt=NUM_FMT, sublabel="PCS")
    _write_kpi_card(ws_d, 4, 6, "누적 입고", total["recv"], fmt=NUM_FMT, sublabel="PCS")
    _write_kpi_card(ws_d, 4, 8, "입고율", total["recv_rate"], fmt=PCT_FMT)
    _write_kpi_card(ws_d, 4, 10, "적기입고율", total["ontime_rate"], fmt=PCT_FMT)
    _write_kpi_card(ws_d, 4, 12, "불량률", total["def_rate"], fmt=PCT2_FMT)

    # 경고 배지
    if total["unreceived"] > 0:
        c = ws_d.cell(row=4, column=14, value=f"미입고 {total['unreceived']} ST")
        c.fill = ALERT_FILL
        c.font = ALERT_FONT
        c.alignment = CENTER
    if total["delay"] > 0:
        c = ws_d.cell(row=5, column=14, value=f"지연 {total['delay']} ST")
        c.fill = RED_FILL
        c.font = Font(bold=True)
        c.alignment = CENTER

    # ── Section 2: 복종별 현황 ──
    sec = 8
    ws_d.cell(row=sec, column=2, value="1. 복종별 발주·입고·적기 현황").font = SEC_FONT
    hdrs = ["복종", "ST", "기획 PCS", "입고 PCS", "입고율", "적기", "지연", "미입고", "적기율", "제조금액", "공급금액"]
    _write_header_row(ws_d, sec + 1, 2, hdrs)

    by_div = _group_by(records, "div")
    r = sec + 2
    for dn in ["OUTER", "TOP", "BOTTOM"]:
        if dn not in by_div:
            continue
        a = _agg(by_div[dn])
        ws_d.cell(row=r, column=2, value=dn).font = BOLD
        ws_d.cell(row=r, column=3, value=a["st"])
        ws_d.cell(row=r, column=4, value=a["plan"]).number_format = NUM_FMT
        ws_d.cell(row=r, column=5, value=a["recv"]).number_format = NUM_FMT
        ws_d.cell(row=r, column=6, value=a["recv_rate"]).number_format = PCT_FMT
        ws_d.cell(row=r, column=7, value=a["ontime"])
        ws_d.cell(row=r, column=8, value=a["delay"])
        ws_d.cell(row=r, column=9, value=a["unreceived"])
        ws_d.cell(row=r, column=10, value=a["ontime_rate"]).number_format = PCT_FMT
        ws_d.cell(row=r, column=11, value=a["mfg_amt"]).number_format = NUM_FMT
        ws_d.cell(row=r, column=12, value=a["supply_amt"]).number_format = NUM_FMT
        # 입고율 조건부 서식
        if a["recv_rate"] < 0.5:
            ws_d.cell(row=r, column=6).fill = RED_FILL
        elif a["recv_rate"] < 0.8:
            ws_d.cell(row=r, column=6).fill = YELLOW_FILL
        else:
            ws_d.cell(row=r, column=6).fill = GREEN_FILL
        r += 1

    # 합계
    ws_d.cell(row=r, column=2, value="합계").font = BOLD
    ws_d.cell(row=r, column=3, value=total["st"])
    ws_d.cell(row=r, column=4, value=total["plan"]).number_format = NUM_FMT
    ws_d.cell(row=r, column=5, value=total["recv"]).number_format = NUM_FMT
    ws_d.cell(row=r, column=6, value=total["recv_rate"]).number_format = PCT_FMT
    ws_d.cell(row=r, column=7, value=total["ontime"])
    ws_d.cell(row=r, column=8, value=total["delay"])
    ws_d.cell(row=r, column=9, value=total["unreceived"])
    ws_d.cell(row=r, column=10, value=total["ontime_rate"]).number_format = PCT_FMT
    ws_d.cell(row=r, column=11, value=total["mfg_amt"]).number_format = NUM_FMT
    ws_d.cell(row=r, column=12, value=total["supply_amt"]).number_format = NUM_FMT

    # ── Section 3: 업체별 성과 히트맵 ──
    sec3 = r + 2
    ws_d.cell(row=sec3, column=2, value="2. 업체별 성과 히트맵").font = SEC_FONT
    hdrs3 = ["업체", "ST", "기획 PCS", "입고 PCS", "입고율", "적기율", "불량률", "미입고", "평균소요일"]
    _write_header_row(ws_d, sec3 + 1, 2, hdrs3)

    by_sup = _group_by(records, "supplier")
    sup_sorted = sorted(by_sup.items(), key=lambda x: len(x[1]), reverse=True)
    r = sec3 + 2
    for sup, recs in sup_sorted:
        if not sup:
            continue
        a = _agg(recs)
        leads = [rc["lead"] for rc in recs if rc["lead"] > 0]
        avg_lead = sum(leads) / len(leads) if leads else 0

        ws_d.cell(row=r, column=2, value=sup)
        ws_d.cell(row=r, column=3, value=a["st"])
        ws_d.cell(row=r, column=4, value=a["plan"]).number_format = NUM_FMT
        ws_d.cell(row=r, column=5, value=a["recv"]).number_format = NUM_FMT
        c_rate = ws_d.cell(row=r, column=6, value=a["recv_rate"])
        c_rate.number_format = PCT_FMT
        c_ot = ws_d.cell(row=r, column=7, value=a["ontime_rate"])
        c_ot.number_format = PCT_FMT
        c_def = ws_d.cell(row=r, column=8, value=a["def_rate"])
        c_def.number_format = PCT2_FMT
        ws_d.cell(row=r, column=9, value=a["unreceived"])
        ws_d.cell(row=r, column=10, value=round(avg_lead, 1) if avg_lead else "-")

        # 히트맵 색상
        if a["recv_rate"] < 0.5:
            c_rate.fill = RED_FILL
        elif a["recv_rate"] < 0.8:
            c_rate.fill = YELLOW_FILL
        else:
            c_rate.fill = GREEN_FILL

        if a["ontime_rate"] < 0.5:
            c_ot.fill = RED_FILL
        elif a["ontime_rate"] < 0.8:
            c_ot.fill = YELLOW_FILL
        else:
            c_ot.fill = GREEN_FILL

        if a["def_rate"] > 0.05:
            c_def.fill = RED_FILL
        elif a["def_rate"] > 0.02:
            c_def.fill = YELLOW_FILL

        r += 1

    # ── Section 4: 생산공정 진행률 히트맵 ──
    sec4 = r + 1
    ws_d.cell(row=sec4, column=2, value="3. 생산공정 진행률 (복종×공정)").font = SEC_FONT
    proc_names = list(PROCESS_COLS.keys())
    _write_header_row(ws_d, sec4 + 1, 2, ["복종"] + proc_names)

    r = sec4 + 2
    for dn in ["OUTER", "TOP", "BOTTOM"]:
        if dn not in by_div:
            continue
        recs = by_div[dn]
        ws_d.cell(row=r, column=2, value=dn).font = BOLD
        for i, pn in enumerate(proc_names):
            done = sum(1 for rc in recs if rc.get(f"proc_{pn}"))
            rate = done / len(recs) if recs else 0
            c = ws_d.cell(row=r, column=3 + i, value=rate)
            c.number_format = PCT_FMT
            c.alignment = CENTER
            if rate >= 0.9:
                c.fill = GREEN_FILL
            elif rate >= 0.5:
                c.fill = YELLOW_FILL
            elif rate > 0:
                c.fill = RED_FILL
        r += 1

    # ── Section 5: 스케줄 진행률 ──
    sec5 = r + 1
    ws_d.cell(row=sec5, column=2, value="4. 스케줄 진행률 (복종×단계)").font = SEC_FONT
    sched_names = list(SCHED_COLS.keys())
    _write_header_row(ws_d, sec5 + 1, 2, ["복종"] + sched_names)

    r = sec5 + 2
    for dn in ["OUTER", "TOP", "BOTTOM"]:
        if dn not in by_div:
            continue
        recs = by_div[dn]
        ws_d.cell(row=r, column=2, value=dn).font = BOLD
        for i, sn in enumerate(sched_names):
            done = sum(1 for rc in recs if rc.get(f"sched_{sn}"))
            rate = done / len(recs) if recs else 0
            c = ws_d.cell(row=r, column=3 + i, value=rate)
            c.number_format = PCT_FMT
            c.alignment = CENTER
            if rate >= 0.9:
                c.fill = GREEN_FILL
            elif rate >= 0.5:
                c.fill = YELLOW_FILL
            elif rate > 0:
                c.fill = RED_FILL
        r += 1

    # ── Section 6: 생산국별 현황 ──
    sec6 = r + 1
    ws_d.cell(row=sec6, column=2, value="5. 생산국별 현황").font = SEC_FONT
    hdrs6 = ["생산국", "ST", "기획 PCS", "입고율", "적기율", "미입고"]
    _write_header_row(ws_d, sec6 + 1, 2, hdrs6)

    by_country = _group_by(records, "country")
    r = sec6 + 2
    for cn, recs in sorted(by_country.items(), key=lambda x: len(x[1]), reverse=True):
        if not cn:
            continue
        a = _agg(recs)
        ws_d.cell(row=r, column=2, value=cn)
        ws_d.cell(row=r, column=3, value=a["st"])
        ws_d.cell(row=r, column=4, value=a["plan"]).number_format = NUM_FMT
        ws_d.cell(row=r, column=5, value=a["recv_rate"]).number_format = PCT_FMT
        ws_d.cell(row=r, column=6, value=a["ontime_rate"]).number_format = PCT_FMT
        ws_d.cell(row=r, column=7, value=a["unreceived"])
        r += 1

    # ── Section 7: 소싱담당별 현황 ──
    sec7 = r + 1
    ws_d.cell(row=sec7, column=2, value="6. 소싱담당별 현황").font = SEC_FONT
    hdrs7 = ["담당자", "ST", "기획 PCS", "입고율", "적기율", "미입고", "지연"]
    _write_header_row(ws_d, sec7 + 1, 2, hdrs7)

    by_sourcer = _group_by(records, "sourcer")
    r = sec7 + 2
    for sc, recs in sorted(by_sourcer.items(), key=lambda x: len(x[1]), reverse=True):
        if not sc:
            continue
        a = _agg(recs)
        ws_d.cell(row=r, column=2, value=sc)
        ws_d.cell(row=r, column=3, value=a["st"])
        ws_d.cell(row=r, column=4, value=a["plan"]).number_format = NUM_FMT
        ws_d.cell(row=r, column=5, value=a["recv_rate"]).number_format = PCT_FMT
        ws_d.cell(row=r, column=6, value=a["ontime_rate"]).number_format = PCT_FMT
        ws_d.cell(row=r, column=7, value=a["unreceived"])
        ws_d.cell(row=r, column=8, value=a["delay"])
        r += 1

    # ── Section 8: 입고 지연 TOP 15 ──
    sec8 = r + 1
    ws_d.cell(row=sec8, column=2, value="7. 입고 지연 TOP 15 (소요일 내림차순)").font = SEC_FONT
    hdrs8 = ["STYLE NO", "COLOR", "복종", "아이템", "업체", "생산국", "확정납기", "소요일", "상태"]
    _write_header_row(ws_d, sec8 + 1, 2, hdrs8)

    delays = [r for r in records if r["ontime"] in ("지연", "지연(7일내)", "미입고")]
    delays.sort(key=lambda x: x["lead"], reverse=True)
    r = sec8 + 2
    for item in delays[:15]:
        ws_d.cell(row=r, column=2, value=item["style"])
        ws_d.cell(row=r, column=3, value=item["color"])
        ws_d.cell(row=r, column=4, value=item["div"])
        ws_d.cell(row=r, column=5, value=item["item"])
        ws_d.cell(row=r, column=6, value=item["supplier"])
        ws_d.cell(row=r, column=7, value=item["country"])
        if item["conf_date"]:
            ws_d.cell(row=r, column=8, value=item["conf_date"]).number_format = DATE_FMT
        ws_d.cell(row=r, column=9, value=int(item["lead"]))
        c_status = ws_d.cell(row=r, column=10, value=item["ontime"])
        c_status.fill = _ontime_fill(item["ontime"])
        r += 1

    # ── Section 9: 원가 분석 (복종×아이템) ──
    sec9 = r + 1
    ws_d.cell(row=sec9, column=2, value="8. 원가 분석 (복종×아이템)").font = SEC_FONT
    hdrs9 = ["복종", "아이템", "ST", "평균사입가", "평균판매가", "평균M/UP", "평균원가율", "총 제조금액", "총 공급금액"]
    _write_header_row(ws_d, sec9 + 1, 2, hdrs9)

    by_div_item = {}
    for rec in records:
        k = (rec["div"], rec["item"])
        if k not in by_div_item:
            by_div_item[k] = []
        by_div_item[k].append(rec)

    r = sec9 + 2
    for (div, item), recs in sorted(by_div_item.items()):
        costs = [rc["cost"] for rc in recs if rc["cost"] > 0]
        prices = [rc["price"] for rc in recs if rc["price"] > 0]
        markups = [rc["markup"] for rc in recs if rc["markup"] > 0]
        cost_rates = [rc["cost_rate"] for rc in recs if rc["cost_rate"] > 0]
        a = _agg(recs)

        ws_d.cell(row=r, column=2, value=div)
        ws_d.cell(row=r, column=3, value=item)
        ws_d.cell(row=r, column=4, value=a["st"])
        ws_d.cell(row=r, column=5, value=round(sum(costs) / len(costs)) if costs else 0).number_format = NUM_FMT
        ws_d.cell(row=r, column=6, value=round(sum(prices) / len(prices)) if prices else 0).number_format = NUM_FMT
        ws_d.cell(row=r, column=7, value=round(sum(markups) / len(markups), 1) if markups else 0)
        ws_d.cell(row=r, column=8, value=sum(cost_rates) / len(cost_rates) if cost_rates else 0).number_format = PCT_FMT
        ws_d.cell(row=r, column=9, value=a["mfg_amt"]).number_format = NUM_FMT
        ws_d.cell(row=r, column=10, value=a["supply_amt"]).number_format = NUM_FMT
        r += 1

    print(f"  종합현황판: {total['st']} ST / 입고율 {total['recv_rate']:.1%} / "
          f"적기율 {total['ontime_rate']:.1%} / 미입고 {total['unreceived']}건")
    return records


# ══════════════════════════════════════════════════════════════
#  주간보고 — 사업부 공유 대시보드
# ══════════════════════════════════════════════════════════════

def 주간보고갱신(wb, records=None):
    ws = ensure_sheet(wb, WS_WEEKLY)
    ws_m = wb[WS_MASTER]
    _clear_sheet(ws)

    if records is None:
        records = _collect_data(ws_m)

    total = _agg(records)
    today = date.today()
    week_num = today.isocalendar()[1]

    # ── 타이틀 ──
    ws.merge_cells("B2:L2")
    ws.cell(row=2, column=2,
            value=f"Wacky Willy 26SS 생산관리 주간보고 | {today:%Y.%m.%d} ({week_num}주차)").font = TITLE_FONT

    # ── A. 주간 핵심 지표 ──
    sec_a = 4
    ws.cell(row=sec_a, column=2, value="A. 주간 핵심 지표").font = SEC_FONT
    _write_kpi_card(ws, sec_a + 1, 2, "총 발주 ST", total["st"])
    _write_kpi_card(ws, sec_a + 1, 4, "기획 PCS", total["plan"], fmt=NUM_FMT)
    _write_kpi_card(ws, sec_a + 1, 6, "누적 입고 PCS", total["recv"], fmt=NUM_FMT)
    _write_kpi_card(ws, sec_a + 1, 8, "입고율", total["recv_rate"], fmt=PCT_FMT)
    _write_kpi_card(ws, sec_a + 1, 10, "적기입고율", total["ontime_rate"], fmt=PCT_FMT)
    _write_kpi_card(ws, sec_a + 1, 12, "불량률", total["def_rate"], fmt=PCT2_FMT)

    # ── B. 복종별 요약 ──
    sec_b = sec_a + 5
    ws.cell(row=sec_b, column=2, value="B. 복종별 현황").font = SEC_FONT
    hdrs_b = ["복종", "ST", "기획 PCS", "입고 PCS", "입고율", "적기", "지연", "미입고",
              "총 제조금액", "총 공급금액"]
    _write_header_row(ws, sec_b + 1, 2, hdrs_b)

    by_div = _group_by(records, "div")
    r = sec_b + 2
    for dn in ["OUTER", "TOP", "BOTTOM"]:
        if dn not in by_div:
            continue
        a = _agg(by_div[dn])
        ws.cell(row=r, column=2, value=dn).font = BOLD
        ws.cell(row=r, column=3, value=a["st"])
        ws.cell(row=r, column=4, value=a["plan"]).number_format = NUM_FMT
        ws.cell(row=r, column=5, value=a["recv"]).number_format = NUM_FMT
        c = ws.cell(row=r, column=6, value=a["recv_rate"])
        c.number_format = PCT_FMT
        if a["recv_rate"] < 0.5:
            c.fill = RED_FILL
        elif a["recv_rate"] < 0.8:
            c.fill = YELLOW_FILL
        else:
            c.fill = GREEN_FILL
        ws.cell(row=r, column=7, value=a["ontime"])
        ws.cell(row=r, column=8, value=a["delay"])
        ws.cell(row=r, column=9, value=a["unreceived"])
        ws.cell(row=r, column=10, value=a["mfg_amt"]).number_format = NUM_FMT
        ws.cell(row=r, column=11, value=a["supply_amt"]).number_format = NUM_FMT
        r += 1

    # 합계
    ws.cell(row=r, column=2, value="합계").font = BOLD
    ws.cell(row=r, column=3, value=total["st"])
    ws.cell(row=r, column=4, value=total["plan"]).number_format = NUM_FMT
    ws.cell(row=r, column=5, value=total["recv"]).number_format = NUM_FMT
    ws.cell(row=r, column=6, value=total["recv_rate"]).number_format = PCT_FMT
    ws.cell(row=r, column=10, value=total["mfg_amt"]).number_format = NUM_FMT
    ws.cell(row=r, column=11, value=total["supply_amt"]).number_format = NUM_FMT

    # ── C. 업체별 입고 성과 ──
    sec_c = r + 2
    ws.cell(row=sec_c, column=2, value="C. 업체별 입고 성과").font = SEC_FONT
    hdrs_c = ["업체", "ST", "기획 PCS", "입고 PCS", "입고율", "적기율", "미입고", "불량률"]
    _write_header_row(ws, sec_c + 1, 2, hdrs_c)

    by_sup = _group_by(records, "supplier")
    r = sec_c + 2
    for sup, recs in sorted(by_sup.items(), key=lambda x: len(x[1]), reverse=True):
        if not sup:
            continue
        a = _agg(recs)
        ws.cell(row=r, column=2, value=sup)
        ws.cell(row=r, column=3, value=a["st"])
        ws.cell(row=r, column=4, value=a["plan"]).number_format = NUM_FMT
        ws.cell(row=r, column=5, value=a["recv"]).number_format = NUM_FMT
        ws.cell(row=r, column=6, value=a["recv_rate"]).number_format = PCT_FMT
        ws.cell(row=r, column=7, value=a["ontime_rate"]).number_format = PCT_FMT
        ws.cell(row=r, column=8, value=a["unreceived"])
        ws.cell(row=r, column=9, value=a["def_rate"]).number_format = PCT2_FMT
        r += 1

    # ── D. 아이템별 진행 현황 ──
    sec_d = r + 1
    ws.cell(row=sec_d, column=2, value="D. 아이템별 진행 현황").font = SEC_FONT
    hdrs_d = ["아이템", "ST", "기획 PCS", "입고율", "BT CF", "메인 CF", "QC AP", "PP", "재단", "봉제", "완성", "수납"]
    _write_header_row(ws, sec_d + 1, 2, hdrs_d)

    by_item = _group_by(records, "item")
    r = sec_d + 2
    for item, recs in sorted(by_item.items(), key=lambda x: len(x[1]), reverse=True):
        if not item:
            continue
        a = _agg(recs)
        ws.cell(row=r, column=2, value=item)
        ws.cell(row=r, column=3, value=a["st"])
        ws.cell(row=r, column=4, value=a["plan"]).number_format = NUM_FMT
        ws.cell(row=r, column=5, value=a["recv_rate"]).number_format = PCT_FMT

        # 스케줄 + 공정 진행률
        proc_checks = [
            ("sched_BT CF", 6), ("sched_메인 CF", 7), ("sched_QC AP", 8), ("sched_PP", 9),
            ("proc_재단", 10), ("proc_봉제", 11), ("proc_완성", 12), ("proc_수납", 13),
        ]
        for key, col in proc_checks:
            done = sum(1 for rc in recs if rc.get(key))
            rate = done / len(recs) if recs else 0
            c = ws.cell(row=r, column=col, value=rate)
            c.number_format = PCT_FMT
            c.alignment = CENTER
            if rate >= 0.9:
                c.fill = GREEN_FILL
            elif rate >= 0.5:
                c.fill = YELLOW_FILL
            elif rate > 0:
                c.fill = RED_FILL
        r += 1

    # ── E. 주요 이슈 (지연 TOP 10) ──
    sec_e = r + 1
    ws.cell(row=sec_e, column=2, value="E. 주요 이슈 — 입고 지연 TOP 10").font = SEC_FONT
    hdrs_e = ["STYLE NO", "COLOR", "복종", "아이템", "업체", "확정납기", "소요일", "상태"]
    _write_header_row(ws, sec_e + 1, 2, hdrs_e)

    delays = sorted([r for r in records if r["ontime"] in ("지연", "지연(7일내)", "미입고")],
                     key=lambda x: x["lead"], reverse=True)
    r = sec_e + 2
    for item in delays[:10]:
        ws.cell(row=r, column=2, value=item["style"])
        ws.cell(row=r, column=3, value=item["color"])
        ws.cell(row=r, column=4, value=item["div"])
        ws.cell(row=r, column=5, value=item["item"])
        ws.cell(row=r, column=6, value=item["supplier"])
        if item["conf_date"]:
            ws.cell(row=r, column=7, value=item["conf_date"]).number_format = DATE_FMT
        ws.cell(row=r, column=8, value=int(item["lead"]))
        c_st = ws.cell(row=r, column=9, value=item["ontime"])
        c_st.fill = _ontime_fill(item["ontime"])
        r += 1

    print(f"  주간보고: {week_num}주차 / {total['st']} ST")


# ══════════════════════════════════════════════════════════════
#  월간보고 — 사업부 공유 대시보드
# ══════════════════════════════════════════════════════════════

def 월간보고갱신(wb, records=None):
    ws = ensure_sheet(wb, WS_MONTHLY)
    ws_m = wb[WS_MASTER]
    _clear_sheet(ws)

    if records is None:
        records = _collect_data(ws_m)

    total = _agg(records)
    today = date.today()

    # ── 타이틀 ──
    ws.merge_cells("B2:L2")
    ws.cell(row=2, column=2,
            value=f"Wacky Willy 26SS 생산관리 월간보고 | {today:%Y년 %m월}").font = TITLE_FONT

    # ── A. 시즌 전체 KPI ──
    sec_a = 4
    ws.cell(row=sec_a, column=2, value="A. 시즌 전체 KPI").font = SEC_FONT
    _write_kpi_card(ws, sec_a + 1, 2, "총 ST", total["st"])
    _write_kpi_card(ws, sec_a + 1, 4, "기획 PCS", total["plan"], fmt=NUM_FMT)
    _write_kpi_card(ws, sec_a + 1, 6, "입고율", total["recv_rate"], fmt=PCT_FMT)
    _write_kpi_card(ws, sec_a + 1, 8, "적기입고율", total["ontime_rate"], fmt=PCT_FMT)
    _write_kpi_card(ws, sec_a + 1, 10, "불량률", total["def_rate"], fmt=PCT2_FMT)
    _write_kpi_card(ws, sec_a + 1, 12, "총 제조금액", total["mfg_amt"], fmt=NUM_FMT)

    # ── B. 월별 입고 실적 (확정납기 월 기준) ──
    sec_b = sec_a + 5
    ws.cell(row=sec_b, column=2, value="B. 월별 입고 실적 (확정납기 기준)").font = SEC_FONT
    hdrs_b = ["월", "ST", "기획 PCS", "입고 PCS", "입고율", "적기", "지연", "미입고", "적기율"]
    _write_header_row(ws, sec_b + 1, 2, hdrs_b)

    by_month = _group_by(records, "conf_month")
    r = sec_b + 2
    for month in sorted(k for k in by_month if k):
        recs = by_month[month]
        a = _agg(recs)
        ws.cell(row=r, column=2, value=month)
        ws.cell(row=r, column=3, value=a["st"])
        ws.cell(row=r, column=4, value=a["plan"]).number_format = NUM_FMT
        ws.cell(row=r, column=5, value=a["recv"]).number_format = NUM_FMT
        c = ws.cell(row=r, column=6, value=a["recv_rate"])
        c.number_format = PCT_FMT
        if a["recv_rate"] < 0.5:
            c.fill = RED_FILL
        elif a["recv_rate"] < 0.8:
            c.fill = YELLOW_FILL
        else:
            c.fill = GREEN_FILL
        ws.cell(row=r, column=7, value=a["ontime"])
        ws.cell(row=r, column=8, value=a["delay"])
        ws.cell(row=r, column=9, value=a["unreceived"])
        ws.cell(row=r, column=10, value=a["ontime_rate"]).number_format = PCT_FMT
        r += 1

    # ── C. 복종×아이템별 실적 ──
    sec_c = r + 1
    ws.cell(row=sec_c, column=2, value="C. 복종×아이템별 실적").font = SEC_FONT
    hdrs_c = ["복종", "아이템", "ST", "기획 PCS", "입고 PCS", "입고율", "적기율", "총 제조금액", "총 공급금액"]
    _write_header_row(ws, sec_c + 1, 2, hdrs_c)

    by_div_item = {}
    for rec in records:
        k = (rec["div"], rec["item"])
        if k not in by_div_item:
            by_div_item[k] = []
        by_div_item[k].append(rec)

    r = sec_c + 2
    for (div, item), recs in sorted(by_div_item.items()):
        a = _agg(recs)
        ws.cell(row=r, column=2, value=div)
        ws.cell(row=r, column=3, value=item)
        ws.cell(row=r, column=4, value=a["st"])
        ws.cell(row=r, column=5, value=a["plan"]).number_format = NUM_FMT
        ws.cell(row=r, column=6, value=a["recv"]).number_format = NUM_FMT
        ws.cell(row=r, column=7, value=a["recv_rate"]).number_format = PCT_FMT
        ws.cell(row=r, column=8, value=a["ontime_rate"]).number_format = PCT_FMT
        ws.cell(row=r, column=9, value=a["mfg_amt"]).number_format = NUM_FMT
        ws.cell(row=r, column=10, value=a["supply_amt"]).number_format = NUM_FMT
        r += 1

    # ── D. 업체별 종합 평가 ──
    sec_d = r + 1
    ws.cell(row=sec_d, column=2, value="D. 업체별 종합 평가").font = SEC_FONT
    hdrs_d = ["업체", "ST", "기획 PCS", "입고율", "적기율", "불량률", "미입고", "평균소요일",
              "총 제조금액"]
    _write_header_row(ws, sec_d + 1, 2, hdrs_d)

    by_sup = _group_by(records, "supplier")
    r = sec_d + 2
    for sup, recs in sorted(by_sup.items(), key=lambda x: len(x[1]), reverse=True):
        if not sup:
            continue
        a = _agg(recs)
        leads = [rc["lead"] for rc in recs if rc["lead"] > 0]
        avg_lead = sum(leads) / len(leads) if leads else 0

        ws.cell(row=r, column=2, value=sup)
        ws.cell(row=r, column=3, value=a["st"])
        ws.cell(row=r, column=4, value=a["plan"]).number_format = NUM_FMT
        ws.cell(row=r, column=5, value=a["recv_rate"]).number_format = PCT_FMT
        ws.cell(row=r, column=6, value=a["ontime_rate"]).number_format = PCT_FMT
        ws.cell(row=r, column=7, value=a["def_rate"]).number_format = PCT2_FMT
        ws.cell(row=r, column=8, value=a["unreceived"])
        ws.cell(row=r, column=9, value=round(avg_lead, 1) if avg_lead else "-")
        ws.cell(row=r, column=10, value=a["mfg_amt"]).number_format = NUM_FMT
        r += 1

    # ── E. 생산국별 실적 ──
    sec_e = r + 1
    ws.cell(row=sec_e, column=2, value="E. 생산국별 실적").font = SEC_FONT
    hdrs_e = ["생산국", "ST", "기획 PCS", "입고 PCS", "입고율", "적기율", "미입고", "총 제조금액"]
    _write_header_row(ws, sec_e + 1, 2, hdrs_e)

    by_country = _group_by(records, "country")
    r = sec_e + 2
    for cn, recs in sorted(by_country.items(), key=lambda x: len(x[1]), reverse=True):
        if not cn:
            continue
        a = _agg(recs)
        ws.cell(row=r, column=2, value=cn)
        ws.cell(row=r, column=3, value=a["st"])
        ws.cell(row=r, column=4, value=a["plan"]).number_format = NUM_FMT
        ws.cell(row=r, column=5, value=a["recv"]).number_format = NUM_FMT
        ws.cell(row=r, column=6, value=a["recv_rate"]).number_format = PCT_FMT
        ws.cell(row=r, column=7, value=a["ontime_rate"]).number_format = PCT_FMT
        ws.cell(row=r, column=8, value=a["unreceived"])
        ws.cell(row=r, column=9, value=a["mfg_amt"]).number_format = NUM_FMT
        r += 1

    # ── F. MAIN vs REORDER 비교 ──
    sec_f = r + 1
    ws.cell(row=sec_f, column=2, value="F. MAIN vs REORDER 비교").font = SEC_FONT
    hdrs_f = ["구분", "ST", "기획 PCS", "입고 PCS", "입고율", "적기율", "총 제조금액", "총 공급금액"]
    _write_header_row(ws, sec_f + 1, 2, hdrs_f)

    by_gubun = _group_by(records, "gubun")
    r = sec_f + 2
    for gn in ["MAIN", "REORDER"]:
        if gn not in by_gubun:
            continue
        a = _agg(by_gubun[gn])
        ws.cell(row=r, column=2, value=gn).font = BOLD
        ws.cell(row=r, column=3, value=a["st"])
        ws.cell(row=r, column=4, value=a["plan"]).number_format = NUM_FMT
        ws.cell(row=r, column=5, value=a["recv"]).number_format = NUM_FMT
        ws.cell(row=r, column=6, value=a["recv_rate"]).number_format = PCT_FMT
        ws.cell(row=r, column=7, value=a["ontime_rate"]).number_format = PCT_FMT
        ws.cell(row=r, column=8, value=a["mfg_amt"]).number_format = NUM_FMT
        ws.cell(row=r, column=9, value=a["supply_amt"]).number_format = NUM_FMT
        r += 1

    print(f"  월간보고: {today:%Y년 %m월} / {total['st']} ST")


# ══════════════════════════════════════════════════════════════
#  메인 — 주간전체갱신
# ══════════════════════════════════════════════════════════════

def 보고서전체갱신(wb):
    """종합현황판 + 주간보고 + 월간보고 갱신 (데이터 1회 수집)."""
    ws_m = wb[WS_MASTER]
    records = _collect_data(ws_m)
    종합현황판갱신(wb, records)
    주간보고갱신(wb, records)
    월간보고갱신(wb, records)


def 주간전체갱신(wb, erp_file=None):
    """Step 1~4 전체 순차 실행."""
    print("\n[Step 1] ERP 데이터 갱신")
    if erp_file:
        erp_갱신_수동(wb, erp_file)
    else:
        erp_갱신_자동(wb)

    print("\n[Step 2] 협력사 데이터 반영")
    협력사데이터반영(wb)

    print("\n[Step 3] 자동 컬럼 계산")
    실입고자동반영(wb)
    불량자동반영(wb)
    적기입고판정(wb)

    print("\n[Step 4] 보고서 갱신")
    보고서전체갱신(wb)


def main():
    parser = argparse.ArgumentParser(description="Wacky Willy 26SS 생산관리 자동화")
    parser.add_argument("--erp", action="store_true", help="ERP 데이터만 갱신")
    parser.add_argument("--supplier", action="store_true", help="협력사 데이터만 반영")
    parser.add_argument("--report", action="store_true", help="보고서만 갱신")
    parser.add_argument("--file", type=str, help="ERP 파일 수동 지정")
    parser.add_argument("--master", type=str, help="메인 파일 경로 (기본: W_26SS_생산관리현황.xlsx)")
    args = parser.parse_args()

    master = Path(args.master) if args.master else MASTER_FILE
    if not master.exists():
        print(f"❌ 메인 파일 없음: {master}")
        sys.exit(1)

    print(f"═══ Wacky Willy 26SS 생산관리 자동화 ═══")
    print(f"파일: {master.name}")
    print(f"시각: {datetime.now():%Y-%m-%d %H:%M:%S}")

    wb = openpyxl.load_workbook(master)

    if not has_sheet(wb, WS_MASTER):
        print(f"❌ '{WS_MASTER}' 시트가 없습니다.")
        wb.close()
        sys.exit(1)

    try:
        if args.erp:
            print("\n[ERP 데이터 갱신]")
            if args.file:
                erp_갱신_수동(wb, args.file)
            else:
                erp_갱신_자동(wb)

        elif args.supplier:
            print("\n[협력사 데이터 반영]")
            협력사데이터반영(wb)

        elif args.report:
            print("\n[보고서 갱신]")
            실입고자동반영(wb)
            불량자동반영(wb)
            적기입고판정(wb)
            보고서전체갱신(wb)

        else:
            주간전체갱신(wb, erp_file=args.file)

        # 저장
        wb.save(master)
        print(f"\n✅ 저장 완료: {master.name}")

    except Exception as e:
        print(f"\n❌ 오류: {e}")
        raise
    finally:
        wb.close()


if __name__ == "__main__":
    main()
