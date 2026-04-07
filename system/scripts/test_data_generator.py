#!/usr/bin/env python3
"""
생산관리 자동화 테스트 데이터 생성기
=====================================
원본 보존 + 가상 ERP / 협력사 / 본사 데이터를 생성하여
production_auto.py 파이프라인을 E2E 테스트.

사용법:
  python test_data_generator.py                # 대화형 메뉴
  python test_data_generator.py --backup       # 원본 백업만
  python test_data_generator.py --erp          # ERP 테스트 데이터 생성
  python test_data_generator.py --supplier     # 협력사 테스트 데이터 생성
  python test_data_generator.py --all          # 전체 생성 + 자동화 실행
  python test_data_generator.py --restore      # 원본 복원
  python test_data_generator.py --clean        # 테스트 데이터 삭제
"""

import argparse
import os
import random
import shutil
import sys
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── 경로 ────────────────────────────────────────────────────
BASE_DIR = Path(__file__).resolve().parent.parent.parent / "workspace" / "26SS" / "sourcing"
MASTER_FILE = BASE_DIR / "W_26SS_생산관리현황.xlsx"
SUPPLIER_DIR = BASE_DIR / "협력사입력"
ERP_DIR = BASE_DIR / "ERP추출"
BACKUP_DIR = BASE_DIR / "_test_backup"

# ── 생산현황 컬럼 (production_auto.py 와 동일) ──────────────
HDR_ROW = 7
DATA_START = 9
C_DIV = 1; C_ITEM = 6; C_STYLENO = 8; C_COLOR = 14
C_COLOR_FULL = 15; C_PLAN_QTY = 25; C_SUPPLIER = 28
C_CONFIRMED = 32; C_COST = 33; C_PRICE = 35

# ERP 원본 컬럼 (헤더 Row 1 기준, SKU2 없음)
ERP_O_ITEM = 7; ERP_O_COLOR = 12; ERP_O_QTY = 50
ERP_R_ITEM = 9; ERP_R_COLOR = 13; ERP_R_QTY = 29; ERP_R_DATE = 1

# ── 스타일 ──────────────────────────────────────────────────
TITLE_FONT = Font(bold=True, size=12)
HDR_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HDR_FONT = Font(bold=True, color="FFFFFF", size=9)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
TEST_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

# ── 가상 데이터 설정 ────────────────────────────────────────
COUNTRIES = ["베트남", "중국", "미얀마", "방글라데시", "인도네시아"]
REGIONS = {
    "베트남": ["호치민", "남딘", "흥엔", "하노이"],
    "중국": ["쯔보", "광저우", "상하이"],
    "미얀마": ["양곤"],
    "방글라데시": ["다카", "치타공"],
    "인도네시아": ["자카르타", "반둥"],
}
FACTORIES = ["Alpha Factory", "Beta Works", "Gamma Textiles", "Delta Sewing"]
LOGISTICS = ["현대물류(3PL)", "CJ대한통운", "한진택배"]
PROCESS_STATUSES = ["원자재입고", "재단", "봉제", "완성", "수납"]


def cell_val(ws, r, c):
    v = ws.cell(row=r, column=c).value
    return v if v is not None else ""


def cell_str(ws, r, c):
    return str(cell_val(ws, r, c)).strip()


def cell_num(ws, r, c):
    v = cell_val(ws, r, c)
    if v is None or v == "":
        return 0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0


def last_row(ws, col):
    for r in range(ws.max_row, 0, -1):
        if ws.cell(row=r, column=col).value is not None:
            return r
    return 0


# ══════════════════════════════════════════════════════════════
#  마스터 시트에서 스타일 정보 수집
# ══════════════════════════════════════════════════════════════

def collect_styles(ws):
    """생산현황에서 테스트에 사용할 스타일 목록 수집."""
    styles = []
    lr = last_row(ws, C_STYLENO)

    for r in range(DATA_START, lr + 1):
        sn = cell_str(ws, r, C_STYLENO)
        cc = cell_str(ws, r, C_COLOR)
        if not sn or not cc:
            continue

        styles.append({
            "row": r,
            "style": sn,
            "color": cc,
            "color_full": cell_str(ws, r, C_COLOR_FULL),
            "div": cell_str(ws, r, C_DIV),
            "item": cell_str(ws, r, C_ITEM),
            "supplier": cell_str(ws, r, C_SUPPLIER),
            "plan_qty": cell_num(ws, r, C_PLAN_QTY),
            "cost": cell_num(ws, r, C_COST),
            "price": cell_num(ws, r, C_PRICE),
            "conf_date": cell_val(ws, r, C_CONFIRMED),
        })

    return styles


# ══════════════════════════════════════════════════════════════
#  Step 0: 백업 / 복원
# ══════════════════════════════════════════════════════════════

def backup_originals():
    """원본 파일 백업."""
    BACKUP_DIR.mkdir(exist_ok=True)

    files_to_backup = [
        MASTER_FILE,
    ]
    # ERP 파일들
    for f in ERP_DIR.glob("erp_data_*.xlsx"):
        files_to_backup.append(f)
    # 협력사 파일들 (템플릿 제외, archive 포함)
    for f in SUPPLIER_DIR.glob("협력사입력_*.xlsx"):
        if "~$" not in f.name:
            files_to_backup.append(f)

    backed = 0
    for fp in files_to_backup:
        if fp.exists():
            rel = fp.relative_to(BASE_DIR)
            dest = BACKUP_DIR / rel
            dest.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(fp, dest)
            backed += 1
            print(f"  ✓ 백업: {rel}")

    # 백업 시점 기록
    (BACKUP_DIR / "_backup_info.txt").write_text(
        f"백업 시점: {datetime.now():%Y-%m-%d %H:%M:%S}\n"
        f"파일 수: {backed}\n"
    )
    print(f"\n  백업 완료: {backed}개 → {BACKUP_DIR}")
    return backed


def restore_originals():
    """백업에서 원본 복원."""
    if not BACKUP_DIR.exists():
        print("  ⚠ 백업 폴더 없음. 먼저 --backup 실행 필요")
        return False

    info = BACKUP_DIR / "_backup_info.txt"
    if info.exists():
        print(f"  {info.read_text().strip()}")

    restored = 0
    for fp in BACKUP_DIR.rglob("*.xlsx"):
        rel = fp.relative_to(BACKUP_DIR)
        dest = BASE_DIR / rel
        dest.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(fp, dest)
        restored += 1
        print(f"  ✓ 복원: {rel}")

    print(f"\n  복원 완료: {restored}개")
    return True


def clean_test_data():
    """테스트 데이터 파일 삭제."""
    deleted = 0

    # ERP 테스트 파일
    for f in ERP_DIR.glob("erp_data_*_TEST.xlsx"):
        f.unlink()
        deleted += 1
        print(f"  ✗ 삭제: {f.name}")

    # 협력사 테스트 파일
    for f in SUPPLIER_DIR.glob("협력사입력_TEST_*.xlsx"):
        f.unlink()
        deleted += 1
        print(f"  ✗ 삭제: {f.name}")

    print(f"\n  삭제 완료: {deleted}개")
    return deleted


# ══════════════════════════════════════════════════════════════
#  Step 1: ERP 테스트 데이터 생성
# ══════════════════════════════════════════════════════════════

def generate_erp_data(styles, sample_pct=0.6):
    """가상 ERP 데이터 (발주데이터 + 입고데이터) 생성.

    Args:
        styles: 마스터 시트에서 수집한 스타일 목록
        sample_pct: 전체 스타일 중 테스트 데이터에 포함할 비율
    """
    today = date.today()
    filename = f"erp_data_{today:%Y.%m.%d}_TEST.xlsx"
    filepath = ERP_DIR / filename

    wb = openpyxl.Workbook()

    # ── 발주데이터 시트 ──
    ws_o = wb.active
    ws_o.title = "발주데이터"

    # 헤더 (ERP 원본 형식 - 86 컬럼)
    order_headers = [
        "발주일자", "딜리버리일", "딜리버리차수", "시즌", "매입처명",
        "납기일자", "품번", "품명", "차수", "발주번호", "사진", "색상",
        "진행상태", "사이즈구분명", "작지확정일", "가격확정일", "납기완료일",
        "비고", "최초등록자", "최초등록일", "최종등록자", "최종등록일",
    ]
    # 사이즈 컬럼 (col 23~50)
    size_names = [
        "F\n170\nF", "WS\n180\nS", "WM\n190\nM", "WL\n200\nL",
        "XS\n210\nXL", "S\n220\n052", "M\n225\n054", "L\n230\n056",
        "XL\n235\n160", "XXL\n240\n190", "24\n245\n220", "25\n250\n",
        "26\n255\n", "27\n260\n", "28\n265\n", "29\n270\n",
        "30\n275\n", "31\n280\n", "32\n285\n", "33\n290\n",
        "34\nWS\n", "35\nS\n", "36\nM\n", "38\nL\n", "40\n\n",
        "42\n\n", "44\n\n",
    ]
    order_headers += size_names
    order_headers += ["발주수량", "발주금액", "사전원가", "사전원가VAT", "예상판매가"]

    for c, h in enumerate(order_headers, 1):
        cell = ws_o.cell(row=1, column=c, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL

    # 데이터 생성
    sampled = random.sample(styles, max(1, int(len(styles) * sample_pct)))
    row = 2

    for s in sampled:
        conf = s["conf_date"]
        if hasattr(conf, "date"):
            order_date = conf - timedelta(days=random.randint(30, 90))
        else:
            order_date = today - timedelta(days=random.randint(30, 120))

        qty = int(s["plan_qty"]) if s["plan_qty"] > 0 else random.randint(500, 3000)
        cost = int(s["cost"]) if s["cost"] > 0 else random.randint(15000, 45000)
        price = int(s["price"]) if s["price"] > 0 else random.randint(59000, 199000)

        ws_o.cell(row=row, column=1, value=order_date.strftime("%Y-%m-%d") if hasattr(order_date, "strftime") else str(order_date))
        ws_o.cell(row=row, column=4, value="26SS")
        ws_o.cell(row=row, column=5, value=s["supplier"])
        ws_o.cell(row=row, column=7, value=s["style"])  # 품번
        ws_o.cell(row=row, column=8, value=f"TEST-{s['item']}")  # 품명
        ws_o.cell(row=row, column=9, value="01")  # 차수
        ws_o.cell(row=row, column=12, value=s["color"])  # 색상
        ws_o.cell(row=row, column=13, value="발주확정")
        ws_o.cell(row=row, column=14, value="의류")

        # 사이즈별 수량 분배 (S/M/L/XL)
        sizes = _distribute_sizes(qty)
        ws_o.cell(row=row, column=29, value=sizes.get("S", 0))
        ws_o.cell(row=row, column=30, value=sizes.get("M", 0))
        ws_o.cell(row=row, column=31, value=sizes.get("L", 0))
        ws_o.cell(row=row, column=32, value=sizes.get("XL", 0))

        ws_o.cell(row=row, column=50, value=qty)  # 발주수량
        ws_o.cell(row=row, column=51, value=qty * cost)  # 발주금액
        ws_o.cell(row=row, column=52, value=cost)  # 사전원가

        row += 1

    print(f"  발주데이터: {row - 2}건")

    # ── 입고데이터 시트 ──
    ws_r = wb.create_sheet(title="입고데이터")

    recv_headers = [
        "입고일자", "입고처명", "입고번호", "비고", "작업구분",
        "물류명", "경리결재", "최초입고일", "품번", "품명", "차수",
        "발주번호", "색상", "진행구분", "매입구분명", "최초판매가",
        "입고단가", "시점금액(현재가*수량)", "사전원가", "사전원가(V+)",
        "수입단가", "사전원가금액", "수입금액", "입고금액", "최초판매금액",
        "성별구분", "성별구분명", "예정계", "계",
    ]
    recv_headers += size_names

    for c, h in enumerate(recv_headers, 1):
        cell = ws_r.cell(row=1, column=c, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL

    # 입고 데이터: 발주 스타일 중 일부만 입고 완료 (다양한 입고율 시뮬레이션)
    row = 2
    recv_count = 0

    for s in sampled:
        # 70% 확률로 입고 발생
        if random.random() > 0.70:
            continue

        qty = int(s["plan_qty"]) if s["plan_qty"] > 0 else random.randint(500, 3000)
        cost = int(s["cost"]) if s["cost"] > 0 else random.randint(15000, 45000)
        price = int(s["price"]) if s["price"] > 0 else random.randint(59000, 199000)

        # 입고율 시뮬레이션: 30~110% (과입고 포함)
        recv_rate = random.uniform(0.3, 1.1)
        recv_qty = int(qty * recv_rate)

        conf = s["conf_date"]
        if hasattr(conf, "date"):
            base_date = conf if isinstance(conf, date) else conf.date()
        else:
            base_date = today - timedelta(days=random.randint(10, 60))

        # 입고일: 확정납기 기준 -7 ~ +21일
        recv_date = base_date + timedelta(days=random.randint(-7, 21))

        ws_r.cell(row=row, column=1, value=recv_date.strftime("%Y-%m-%d"))
        ws_r.cell(row=row, column=2, value=s["supplier"])
        ws_r.cell(row=row, column=3, value=f"T{recv_count+1:04d}")
        ws_r.cell(row=row, column=5, value="입고")
        ws_r.cell(row=row, column=6, value=random.choice(LOGISTICS))
        ws_r.cell(row=row, column=7, value="N")
        ws_r.cell(row=row, column=8, value=recv_date.strftime("%Y-%m-%d"))
        ws_r.cell(row=row, column=9, value=s["style"])  # 품번
        ws_r.cell(row=row, column=10, value=f"TEST-{s['item']}")
        ws_r.cell(row=row, column=11, value="01")
        ws_r.cell(row=row, column=13, value=s["color"])  # 색상
        ws_r.cell(row=row, column=14, value="진행")
        ws_r.cell(row=row, column=15, value="완사입")
        ws_r.cell(row=row, column=16, value=price)
        ws_r.cell(row=row, column=17, value=cost)
        ws_r.cell(row=row, column=18, value=price * recv_qty)
        ws_r.cell(row=row, column=28, value=recv_qty)  # 예정계
        ws_r.cell(row=row, column=29, value=recv_qty)  # 계

        # 사이즈별 분배
        sizes = _distribute_sizes(recv_qty)
        ws_r.cell(row=row, column=36, value=sizes.get("S", 0))
        ws_r.cell(row=row, column=37, value=sizes.get("M", 0))
        ws_r.cell(row=row, column=38, value=sizes.get("L", 0))
        ws_r.cell(row=row, column=39, value=sizes.get("XL", 0))

        row += 1
        recv_count += 1

    print(f"  입고데이터: {recv_count}건")

    wb.save(filepath)
    print(f"  ✓ 저장: {filepath.name}")
    return filepath


def _distribute_sizes(total_qty):
    """수량을 사이즈별로 분배 (S:20%, M:30%, L:30%, XL:20%)."""
    ratios = {"S": 0.20, "M": 0.30, "L": 0.30, "XL": 0.20}
    sizes = {}
    remaining = total_qty
    for sz, ratio in list(ratios.items())[:-1]:
        q = int(total_qty * ratio)
        sizes[sz] = q
        remaining -= q
    sizes["XL"] = remaining
    return sizes


# ══════════════════════════════════════════════════════════════
#  Step 2: 협력사 테스트 데이터 생성
# ══════════════════════════════════════════════════════════════

def generate_supplier_data(styles, sample_pct=0.4):
    """가상 협력사 입력 데이터 생성 (신규 양식: 단일 시트).

    Args:
        styles: 마스터 시트에서 수집한 스타일 목록
        sample_pct: 전체 스타일 중 테스트 데이터에 포함할 비율
    """
    today = date.today()
    filename = f"협력사입력_TEST_{today:%Y%m%d}.xlsx"
    filepath = SUPPLIER_DIR / filename

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "협력사입력"

    # ── 헤더 (4행 구조, 템플릿과 동일) ──
    ws.cell(row=1, column=1, value="◆ 협력사 종합 입력 양식 — TEST DATA").font = TITLE_FONT
    ws.cell(row=2, column=1, value="※ 테스트 데이터 — 자동 생성").font = Font(color="FF0000")

    # Row 3: 그룹 헤더
    group_headers = {
        1: "키", 3: "CO", 6: "스케줄 — 원자재",
        33: "스케줄 — QC/PP", 53: "생산진행", 62: "입고예정",
    }
    for col, txt in group_headers.items():
        ws.cell(row=3, column=col, value=txt).font = Font(bold=True)

    # Row 4: 상세 헤더
    detail_headers = [
        "STYLE NO", "COLOR",  # 1-2
        "생산국", "생산 지역", "공장 명",  # 3-5 (CO)
        "원단처", "자재명", "혼용률", "작지투입",  # 6-9
        "BT 의뢰", "BT 1차", "회신", "BT 2차", "회신",  # 10-14
        "BT 3차", "회신", "BT 4차", "회신", "BT 5차", "회신",  # 15-20
        "CF", "진도율",  # 21-22
        "메인", "CF", "메인 1차", "회신", "메인 2차", "회신",  # 23-28
        "메인 3차", "회신", "CF", "진도율",  # 29-32
        "성적서\n발행일", "PASS", "SD", "보정",  # 33-36
        "QC 1차", "회신", "QC 2차", "회신", "QC 3차", "회신",  # 37-42
        "AP", "ST", "진도율",  # 43-45
        "PP 1차", "회신", "PP 2차", "회신", "PP", "ST", "진도율",  # 46-52
        "원자재\n입고", "자재\n확정", "자재\n입고", "재단",  # 53-56
        "외주\n(프린트)", "봉제", "외주\n(워싱)", "완성", "수납",  # 57-61
        "SD", "1차\n입고일", "입고\n수량", "2차\n입고일", "입고\n수량",  # 62-66
        "3차\n입고일", "입고\n수량", "4차\n입고일", "입고\n수량",  # 67-70
        "5차\n입고일", "입고\n수량", "6차\n입고일", "입고\n수량", "계",  # 71-75
    ]
    for c, h in enumerate(detail_headers, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CENTER

    # ── 데이터 생성 ──
    sampled = random.sample(styles, max(1, int(len(styles) * sample_pct)))
    row = 5

    for s in sampled:
        conf = s["conf_date"]
        if hasattr(conf, "date"):
            base_date = conf if isinstance(conf, date) else conf.date()
        else:
            base_date = date.today() - timedelta(days=30)

        # col 1-2: 키
        ws.cell(row=row, column=1, value=s["style"])
        ws.cell(row=row, column=2, value=s["color"])

        # col 3-5: CO
        country = random.choice(COUNTRIES)
        region = random.choice(REGIONS[country])
        ws.cell(row=row, column=3, value=country)
        ws.cell(row=row, column=4, value=region)
        ws.cell(row=row, column=5, value=random.choice(FACTORIES))

        # col 6-8: 원단처 정보
        ws.cell(row=row, column=6, value=random.choice(["동방직물", "대한텍스", "삼양직물", "한솔직물"]))
        ws.cell(row=row, column=7, value=random.choice(["T/C 트윌", "폴리 옥스포드", "면 20수", "나일론 립스탑"]))
        ws.cell(row=row, column=8, value=random.choice(["T65/C35", "P100", "C100", "N100"]))

        # 스케줄 — 원자재: 날짜 진행 시뮬레이션
        sched_base = base_date - timedelta(days=random.randint(60, 120))
        progress_level = random.choice(["early", "mid", "late", "complete"])

        if progress_level in ("mid", "late", "complete"):
            ws.cell(row=row, column=9, value=sched_base)  # 작지투입
            ws.cell(row=row, column=10, value=sched_base + timedelta(days=3))  # BT 의뢰
            ws.cell(row=row, column=11, value=sched_base + timedelta(days=10))  # BT 1차
            ws.cell(row=row, column=12, value=sched_base + timedelta(days=14))  # 회신

        if progress_level in ("late", "complete"):
            ws.cell(row=row, column=21, value=sched_base + timedelta(days=30))  # CF
            ws.cell(row=row, column=23, value=sched_base + timedelta(days=35))  # 메인
            ws.cell(row=row, column=25, value=sched_base + timedelta(days=40))  # 메인 1차
            ws.cell(row=row, column=26, value=sched_base + timedelta(days=44))  # 회신
            ws.cell(row=row, column=31, value=sched_base + timedelta(days=50))  # CF
            ws.cell(row=row, column=32, value="100%")  # 진도율

        # 스케줄 — QC/PP
        if progress_level in ("late", "complete"):
            qc_base = sched_base + timedelta(days=55)
            ws.cell(row=row, column=33, value=qc_base)  # 성적서 발행일
            ws.cell(row=row, column=34, value="PASS")
            ws.cell(row=row, column=35, value=qc_base + timedelta(days=3))  # SD
            ws.cell(row=row, column=37, value=qc_base + timedelta(days=7))  # QC 1차
            ws.cell(row=row, column=38, value=qc_base + timedelta(days=10))  # 회신

        if progress_level == "complete":
            ws.cell(row=row, column=43, value=qc_base + timedelta(days=15))  # AP
            ws.cell(row=row, column=44, value=1)  # ST
            ws.cell(row=row, column=45, value="100%")  # 진도율
            ws.cell(row=row, column=46, value=qc_base + timedelta(days=18))  # PP 1차
            ws.cell(row=row, column=47, value=qc_base + timedelta(days=21))  # 회신
            ws.cell(row=row, column=50, value=qc_base + timedelta(days=25))  # PP
            ws.cell(row=row, column=51, value=1)  # ST
            ws.cell(row=row, column=52, value="100%")  # 진도율

        # 생산진행 (col 53-61)
        if progress_level in ("mid", "late", "complete"):
            prod_base = base_date - timedelta(days=random.randint(30, 60))
            ws.cell(row=row, column=53, value=prod_base)  # 원자재입고
            ws.cell(row=row, column=54, value=prod_base + timedelta(days=2))  # 자재확정
            ws.cell(row=row, column=55, value=prod_base + timedelta(days=5))  # 자재입고

        if progress_level in ("late", "complete"):
            ws.cell(row=row, column=56, value=prod_base + timedelta(days=8))  # 재단
            ws.cell(row=row, column=58, value=prod_base + timedelta(days=15))  # 봉제

        if progress_level == "complete":
            ws.cell(row=row, column=60, value=prod_base + timedelta(days=25))  # 완성
            ws.cell(row=row, column=61, value=prod_base + timedelta(days=28))  # 수납

        # 입고예정 (col 62-75)
        qty = int(s["plan_qty"]) if s["plan_qty"] > 0 else random.randint(500, 3000)
        ship_date = base_date - timedelta(days=random.randint(-5, 10))

        # 1차 입고예정
        ws.cell(row=row, column=62, value=ship_date - timedelta(days=7))  # SD
        ws.cell(row=row, column=63, value=ship_date)  # 1차 입고일
        ws.cell(row=row, column=64, value=int(qty * 0.6))  # 1차 수량

        # 2차 입고예정 (60% 확률)
        if random.random() > 0.4:
            ws.cell(row=row, column=65, value=ship_date + timedelta(days=14))  # 2차 입고일
            ws.cell(row=row, column=66, value=int(qty * 0.3))  # 2차 수량

            # 3차 (30% 확률)
            if random.random() > 0.7:
                ws.cell(row=row, column=67, value=ship_date + timedelta(days=28))
                ws.cell(row=row, column=68, value=int(qty * 0.1))

        # 계 (col 75)
        total_exp = sum(
            (ws.cell(row=row, column=c).value or 0)
            for c in [64, 66, 68, 70, 72, 74]
        )
        ws.cell(row=row, column=75, value=total_exp)

        row += 1

    print(f"  협력사입력: {row - 5}건 ({filepath.name})")
    wb.save(filepath)
    print(f"  ✓ 저장: {filepath.name}")
    return filepath


# ══════════════════════════════════════════════════════════════
#  대화형 메뉴
# ══════════════════════════════════════════════════════════════

MENU = """
╔══════════════════════════════════════════════════════╗
║  생산관리 자동화 테스트 데이터 생성기                  ║
╠══════════════════════════════════════════════════════╣
║                                                      ║
║  [1] 원본 백업                                       ║
║  [2] ERP 테스트 데이터 생성                           ║
║  [3] 협력사 테스트 데이터 생성                        ║
║  [4] 전체 생성 + 자동화 실행                          ║
║  [5] 원본 복원                                       ║
║  [6] 테스트 데이터 삭제                               ║
║  [0] 종료                                            ║
║                                                      ║
╚══════════════════════════════════════════════════════╝
"""


def run_automation():
    """production_auto.py 실행."""
    script = Path(__file__).parent / "production_auto.py"
    if not script.exists():
        print(f"  ⚠ 스크립트 없음: {script}")
        return False

    print("\n" + "=" * 50)
    print("  production_auto.py 실행 중...")
    print("=" * 50)

    import subprocess
    result = subprocess.run(
        [sys.executable, str(script)],
        capture_output=False,
        cwd=str(script.parent),
    )
    return result.returncode == 0


def load_master():
    """마스터 파일 로드 + 스타일 수집."""
    if not MASTER_FILE.exists():
        print(f"  ⚠ 마스터 파일 없음: {MASTER_FILE}")
        sys.exit(1)

    wb = openpyxl.load_workbook(MASTER_FILE, data_only=True)
    ws = wb["생산현황"]
    styles = collect_styles(ws)
    wb.close()
    print(f"  마스터 로드: {len(styles)} 스타일")
    return styles


def interactive_menu():
    """대화형 메뉴."""
    print(MENU)

    while True:
        choice = input("  선택 [0-6]: ").strip()

        if choice == "0":
            print("\n  종료합니다.")
            break

        elif choice == "1":
            print("\n── 원본 백업 ──")
            backup_originals()

        elif choice == "2":
            print("\n── ERP 테스트 데이터 생성 ──")
            styles = load_master()
            generate_erp_data(styles)

        elif choice == "3":
            print("\n── 협력사 테스트 데이터 생성 ──")
            styles = load_master()
            generate_supplier_data(styles)

        elif choice == "4":
            print("\n── 전체 테스트 실행 ──")
            print("\n[Step 0] 원본 백업")
            backup_originals()

            print("\n[Step 1] 테스트 데이터 생성")
            styles = load_master()
            erp_path = generate_erp_data(styles)
            sup_path = generate_supplier_data(styles)

            print("\n[Step 2] 자동화 파이프라인 실행")
            run_automation()

            print("\n" + "=" * 50)
            print("  테스트 완료!")
            print(f"  ERP: {erp_path.name}")
            print(f"  협력사: {sup_path.name}")
            print(f"  결과: {MASTER_FILE.name} 확인")
            print("=" * 50)

        elif choice == "5":
            print("\n── 원본 복원 ──")
            restore_originals()

        elif choice == "6":
            print("\n── 테스트 데이터 삭제 ──")
            clean_test_data()

        else:
            print("  ⚠ 올바른 번호를 입력하세요.")

        print()


# ══════════════════════════════════════════════════════════════
#  메인
# ══════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="생산관리 자동화 테스트 데이터 생성기")
    parser.add_argument("--backup", action="store_true", help="원본 백업")
    parser.add_argument("--erp", action="store_true", help="ERP 테스트 데이터 생성")
    parser.add_argument("--supplier", action="store_true", help="협력사 테스트 데이터 생성")
    parser.add_argument("--all", action="store_true", help="전체 생성 + 자동화 실행")
    parser.add_argument("--restore", action="store_true", help="원본 복원")
    parser.add_argument("--clean", action="store_true", help="테스트 데이터 삭제")
    parser.add_argument("--pct", type=float, default=0.5, help="테스트 샘플 비율 (0.0~1.0)")

    args = parser.parse_args()
    has_flag = any([args.backup, args.erp, args.supplier, args.all, args.restore, args.clean])

    if not has_flag:
        interactive_menu()
        return

    if args.backup:
        backup_originals()

    if args.erp:
        styles = load_master()
        generate_erp_data(styles, sample_pct=args.pct)

    if args.supplier:
        styles = load_master()
        generate_supplier_data(styles, sample_pct=args.pct)

    if args.all:
        print("[Step 0] 원본 백업")
        backup_originals()

        print("\n[Step 1] 테스트 데이터 생성")
        styles = load_master()
        generate_erp_data(styles, sample_pct=args.pct)
        generate_supplier_data(styles, sample_pct=args.pct)

        print("\n[Step 2] 자동화 파이프라인 실행")
        run_automation()

    if args.restore:
        restore_originals()

    if args.clean:
        clean_test_data()


if __name__ == "__main__":
    main()
