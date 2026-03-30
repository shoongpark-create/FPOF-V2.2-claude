"""
Musinsa Ranking Crawler for Wacky Willy MD Planning
====================================================
무신사 랭킹 API를 통해 카테고리별/기간별/성별/연령별
상품 랭킹 데이터를 수집하고 엑셀로 저장하는 파이프라인.

Selenium 없이 API 직접 호출 방식 — 빠르고 안정적.

사용법:
  python crawler.py --category 000 --period REALTIME --gender A --age AGE_BAND_ALL
  python crawler.py --category 001 --period WEEKLY --gender F --age AGE_BAND_20
  python crawler.py --category 001,002,003 --period DAILY --gender A --age AGE_BAND_ALL
  python crawler.py --list-categories          # 카테고리 목록 출력
"""

from __future__ import annotations

import os
import sys
import json
import argparse
import requests
from datetime import datetime

from io import BytesIO

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XlImage
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from PIL import Image as PILImage
    HAS_PILLOW = True
except ImportError:
    HAS_PILLOW = False


# ──────────────────────────────────────────────
# 1. 카테고리 & 필터 매핑
# ──────────────────────────────────────────────

CATEGORIES = {
    "000": "전체",
    "001": "상의",
    "002": "아우터",
    "003": "바지",
    "004": "가방",
    "100": "원피스/스커트",
    "101": "소품",
    "103": "신발",
    "104": "뷰티",
    "026": "속옷/홈웨어",
    "017": "스포츠/키즈",
}

# 2depth 서브카테고리 (주요 카테고리만)
SUB_CATEGORIES = {
    "001": {  # 상의
        "001001": "반소매 티셔츠", "001002": "긴소매 티셔츠", "001003": "셔츠/블라우스",
        "001004": "피케/카라 티셔츠", "001005": "맨투맨/스웨트셔츠", "001006": "후드 티셔츠",
        "001010": "니트/스웨터", "001011": "민소매 티셔츠",
    },
    "002": {  # 아우터
        "002001": "후드 집업", "002002": "블루종/MA-1", "002003": "레더/라이더스 재킷",
        "002004": "무스탕/퍼", "002006": "코트", "002007": "롱패딩/헤비 아우터",
        "002008": "숏패딩/헤비 아우터", "002012": "카디건", "002017": "나일론/코치 재킷",
        "002022": "기타 아우터",
    },
    "003": {  # 바지
        "003002": "데님 팬츠", "003004": "코튼 팬츠", "003007": "트레이닝/조거 팬츠",
        "003008": "숏 팬츠", "003009": "슬랙스", "003025": "점프 슈트/오버올",
        "003011": "레깅스",
    },
    "103": {  # 신발
        "103001": "스니커즈", "103002": "구두", "103004": "샌들",
        "103005": "슬리퍼", "103007": "부츠/워커",
    },
}

PERIODS = {
    "REALTIME": "실시간",
    "DAILY": "일간",
    "WEEKLY": "주간",
    "MONTHLY": "월간",
}

GENDERS = {
    "A": "전체",
    "M": "남성",
    "F": "여성",
}

AGE_BANDS = {
    "AGE_BAND_ALL": "전체 연령",
    "AGE_BAND_MINOR": "19세 이하",
    "AGE_BAND_20": "20-24세",
    "AGE_BAND_25": "25-29세",
    "AGE_BAND_30": "30-34세",
    "AGE_BAND_35": "35-39세",
    "AGE_BAND_40": "40세 이상",
}

# sectionId 매핑 (기간별)
SECTION_IDS = {
    "REALTIME": 199,
    "DAILY": 200,
    "WEEKLY": 201,
    "MONTHLY": 202,
}

# ──────────────────────────────────────────────
# 2. API 호출
# ──────────────────────────────────────────────

BASE_URL = "https://api.musinsa.com/api2/hm/web/v5/pans/ranking/sections"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) "
                  "Chrome/131.0.0.0 Safari/537.36",
    "Accept": "application/json",
    "Referer": "https://www.musinsa.com/",
    "Origin": "https://www.musinsa.com",
}


def fetch_ranking(category: str, period: str, gender: str, age_band: str) -> list[dict]:
    """무신사 랭킹 API에서 상품 데이터를 가져온다."""
    section_id = SECTION_IDS.get(period, 200)
    url = f"{BASE_URL}/{section_id}"

    params = {
        "storeCode": "musinsa",
        "categoryCode": category,
        "gf": gender,
        "ageBand": age_band,
    }

    print(f"  API 호출: section={section_id}, category={category}, "
          f"gender={gender}, age={age_band}")

    try:
        resp = requests.get(url, params=params, headers=HEADERS, timeout=30)
        resp.raise_for_status()
        data = resp.json()
    except requests.RequestException as e:
        print(f"  [ERROR] API 호출 실패: {e}")
        return []

    # modules에서 상품 데이터 추출
    # 구조: data.modules[] → type=MULTICOLUMN → items[] → type=PRODUCT_COLUMN
    # 상품 정보는 info/image/onClick 하위에 분산되어 있음
    products = []
    modules = data.get("data", {}).get("modules", [])

    for module in modules:
        mod_type = module.get("type", "")
        if mod_type == "MULTICOLUMN":
            items = module.get("items", [])
            for item in items:
                if item.get("type") == "PRODUCT_COLUMN":
                    product = parse_product(item)
                    if product:
                        products.append(product)

    print(f"  수집 완료: {len(products)}개 상품")
    return products


def parse_product(item: dict) -> dict | None:
    """API 응답의 개별 상품 데이터를 정리한다.

    무신사 API 구조:
      item.info     → brandName, productName, finalPrice, discountRatio, additionalInformation[]
      item.image    → rank, url, labels[] (급상승/판매수 등 뱃지)
      item.onClick  → url (상품 페이지 링크)
    """
    try:
        info = item.get("info", {})
        image = item.get("image", {})
        on_click = item.get("onClick", {})

        # 상품 URL
        product_url = on_click.get("url", "") if isinstance(on_click, dict) else ""

        # 뱃지/라벨 (급상승, 판매 1.5천개 등)
        labels = image.get("labels", [])
        label_texts = [lbl.get("text", "") for lbl in labels if lbl.get("text")]
        label_str = ", ".join(label_texts)

        # 실시간 정보 (N명이 보는 중, N명이 구매 중)
        add_infos = info.get("additionalInformation", [])
        add_info_str = " | ".join(ai.get("text", "") for ai in add_infos if ai.get("text"))

        # 옵션 가격 정보
        option_price = info.get("optionPriceInformation", {})
        option_text = option_price.get("text", "") if isinstance(option_price, dict) else ""

        # 원가 (originalPrice가 없으면 finalPrice와 discountRatio로 역산)
        original_price = info.get("originalPrice", 0)
        final_price = info.get("finalPrice", 0)
        discount_ratio = info.get("discountRatio", 0)
        if not original_price and discount_ratio and final_price:
            original_price = int(final_price / (1 - discount_ratio / 100))

        return {
            "rank": image.get("rank", 0),
            "brand_name": info.get("brandName", ""),
            "product_name": info.get("productName", ""),
            "product_id": item.get("id", ""),
            "original_price": original_price,
            "final_price": final_price,
            "discount_ratio": discount_ratio,
            "best_price": info.get("bestPrice", 0),
            "image_url": image.get("url", ""),
            "product_url": product_url,
            "labels": label_str,                    # "급상승", "판매 1.5천개" 등
            "additional_info": add_info_str,        # "430명이 보는 중 | 38명이 구매 중"
            "option_price": option_text,
            "strikethrough": info.get("strikethrough", False),
        }
    except Exception as e:
        print(f"  [WARN] 상품 파싱 실패: {e}")
        return None


# ──────────────────────────────────────────────
# 3. 엑셀 저장
# ──────────────────────────────────────────────

IMG_COL = 1          # A열에 이미지 삽입
IMG_WIDTH_PX = 60    # 셀 내 이미지 너비 (px)
IMG_HEIGHT_PX = 75   # 셀 내 이미지 높이 (px)
ROW_HEIGHT_PT = 60   # 이미지 행 높이 (pt, ≈ 80px)


def _download_thumbnail(url: str) -> bytes | None:
    """상품 썸네일을 다운로드하여 bytes로 반환한다."""
    if not url:
        return None
    try:
        if url.startswith("//"):
            url = "https:" + url
        # 무신사 API 기본값이 _500.jpg — 그대로 사용 (고화질)
        resp = requests.get(url, headers=HEADERS, timeout=10)
        if resp.status_code == 200 and len(resp.content) > 1000:
            return resp.content
    except Exception:
        pass
    return None


def save_to_excel(products: list[dict], output_path: str, meta: dict, embed_images: bool = True):
    """수집된 상품 데이터를 엑셀 파일로 저장한다.
    embed_images=True이면 A열에 상품 썸네일을 직접 삽입한다.
    """
    if not HAS_OPENPYXL:
        print("  [WARN] openpyxl 미설치 — JSON으로 대체 저장합니다.")
        json_path = output_path.replace(".xlsx", ".json")
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump({"meta": meta, "products": products}, f, ensure_ascii=False, indent=2)
        print(f"  저장: {json_path}")
        return json_path

    wb = Workbook()
    ws = wb.active
    ws.title = "무신사 랭킹"

    # 이미지 삽입 가능 여부
    can_embed = embed_images and HAS_OPENPYXL and HAS_PILLOW

    # 컬럼 정의 — 이미지는 상품명 다음(D열)
    columns = [
        ("순위", 6),          # A
        ("브랜드", 18),       # B
        ("상품명", 40),       # C
        ("이미지", 10),       # D: 썸네일
        ("원가", 12),         # E
        ("할인율", 8),        # F
        ("판매가", 12),       # G
        ("최저가", 12),       # H
        ("뱃지/라벨", 16),    # I
        ("실시간 정보", 24),   # J
        ("옵션가", 12),       # K
        ("상품 URL", 50),     # L
        ("이미지 URL", 55),   # M
    ]
    IMG_COL_IDX = 4  # D열 = 이미지 컬럼

    # 메타 정보 헤더
    last_col_letter = get_column_letter(len(columns))
    meta_text = (f"무신사 랭킹 | "
                 f"카테고리: {meta.get('category_name', '')} | "
                 f"기간: {meta.get('period_name', '')} | "
                 f"성별: {meta.get('gender_name', '')} | "
                 f"연령: {meta.get('age_name', '')} | "
                 f"수집: {meta.get('crawled_at', '')}")
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = meta_text
    ws["A1"].font = Font(bold=True, size=11)
    ws["A1"].fill = PatternFill("solid", fgColor="E8E8E8")

    # 컬럼 헤더
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor="333333")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    for col_idx, (col_name, col_width) in enumerate(columns, 1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # 이미지 다운로드 (embed_images=True일 때)
    image_data_map = {}
    if can_embed:
        print("  이미지 다운로드 중...")
        downloaded = 0
        for i, p in enumerate(products):
            img_bytes = _download_thumbnail(p.get("image_url", ""))
            if img_bytes:
                image_data_map[i] = img_bytes
                downloaded += 1
            # 진행률 (25개마다)
            if (i + 1) % 25 == 0:
                print(f"    {i + 1}/{len(products)} ({downloaded}장 다운로드)")
        print(f"  이미지 다운로드 완료: {downloaded}/{len(products)}장")

    # 데이터 행
    for row_idx, p in enumerate(products, 3):
        data_idx = row_idx - 3

        # A~L열: 텍스트 데이터 (D열=이미지는 빈 셀)
        row_data = [
            p["rank"],                                                    # A: 순위
            p["brand_name"],                                              # B: 브랜드
            p["product_name"],                                            # C: 상품명
            "",                                                           # D: 이미지 (빈 셀 → 나중에 삽입)
            p["original_price"],                                          # E: 원가
            f'{p["discount_ratio"]}%' if p["discount_ratio"] else "",     # F: 할인율
            p["final_price"],                                             # G: 판매가
            p["best_price"] if p["best_price"] else "",                   # H: 최저가
            p.get("labels", ""),                                          # I: 뱃지
            p["additional_info"],                                         # J: 실시간
            p.get("option_price", ""),                                    # K: 옵션가
            p["product_url"],                                             # L: 상품 URL
            p.get("image_url", ""),                                       # M: 이미지 URL
        ]
        # 텍스트 줄바꿈이 필요한 컬럼: B(브랜드), C(상품명), I(뱃지), J(실시간), L(URL)
        WRAP_COLS = {2, 3, 9, 10, 12, 13}

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx in WRAP_COLS:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(vertical="center")
            if col_idx in (5, 7, 8):  # 가격 컬럼 (E, G, H)
                cell.number_format = '#,##0'
            if col_idx in (12, 13):  # URL 컬럼 (L, M)
                cell.font = Font(color="0066CC", underline="single", size=9)
            if col_idx == IMG_COL_IDX:  # 이미지 컬럼 (D)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # 행 높이 설정 (이미지가 들어갈 공간)
        if can_embed:
            ws.row_dimensions[row_idx].height = ROW_HEIGHT_PT

        # 짝수행 배경색
        if row_idx % 2 == 0:
            for col_idx in range(1, len(columns) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = PatternFill("solid", fgColor="F8F8F8")

        # 이미지 삽입
        if can_embed and data_idx in image_data_map:
            try:
                img_bytes = image_data_map[data_idx]
                img_stream = BytesIO(img_bytes)
                pil_img = PILImage.open(img_stream)

                # 비율 유지하며 리사이즈
                pil_img.thumbnail((IMG_WIDTH_PX, IMG_HEIGHT_PX), PILImage.LANCZOS)
                resized_stream = BytesIO()
                pil_img.save(resized_stream, format="JPEG", quality=85)
                resized_stream.seek(0)

                xl_img = XlImage(resized_stream)
                # D열(이미지 컬럼)의 해당 행에 앵커
                img_col_letter = get_column_letter(IMG_COL_IDX)
                cell_ref = f"{img_col_letter}{row_idx}"
                ws.add_image(xl_img, cell_ref)
            except Exception:
                pass  # 이미지 실패해도 데이터는 유지

    # 필터 설정
    ws.auto_filter.ref = f"A2:{last_col_letter}{len(products) + 2}"

    # 틀 고정 (이미지 컬럼 다음부터 스크롤)
    ws.freeze_panes = "E3"

    wb.save(output_path)
    print(f"  저장: {output_path}")
    return output_path


def save_to_json(products: list[dict], output_path: str, meta: dict):
    """수집된 상품 데이터를 JSON 파일로 저장한다."""
    payload = {
        "meta": meta,
        "products": products,
        "count": len(products),
    }
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    print(f"  저장: {output_path}")
    return output_path


# ──────────────────────────────────────────────
# 4. 이미지 다운로드 (선택)
# ──────────────────────────────────────────────

def download_images(products: list[dict], output_dir: str):
    """상품 대표 이미지를 다운로드한다."""
    img_dir = os.path.join(output_dir, "images")
    os.makedirs(img_dir, exist_ok=True)

    downloaded = 0
    for p in products:
        url = p.get("image_url", "")
        if not url:
            continue
        try:
            # URL이 // 로 시작하면 https: 붙이기
            if url.startswith("//"):
                url = "https:" + url
            resp = requests.get(url, headers=HEADERS, timeout=15)
            if resp.status_code == 200 and len(resp.content) > 5000:
                fname = f'{p["rank"]:03d}_{p["brand_name"]}_{p["product_id"]}.jpg'
                # 파일명에서 위험 문자 제거
                fname = "".join(c if c.isalnum() or c in "._-" else "_" for c in fname)
                fpath = os.path.join(img_dir, fname)
                with open(fpath, "wb") as f:
                    f.write(resp.content)
                downloaded += 1
        except Exception:
            continue

    print(f"  이미지 다운로드: {downloaded}/{len(products)}장")
    return downloaded


# ──────────────────────────────────────────────
# 5. CLI
# ──────────────────────────────────────────────

def list_categories():
    """사용 가능한 카테고리 목록을 출력한다."""
    print("\n[카테고리 목록]")
    for code, name in CATEGORIES.items():
        print(f"  {code}: {name}")
        subs = SUB_CATEGORIES.get(code, {})
        for sub_code, sub_name in subs.items():
            print(f"    └─ {sub_code}: {sub_name}")

    print("\n[기간]")
    for code, name in PERIODS.items():
        print(f"  {code}: {name}")

    print("\n[성별]")
    for code, name in GENDERS.items():
        print(f"  {code}: {name}")

    print("\n[연령대]")
    for code, name in AGE_BANDS.items():
        print(f"  {code}: {name}")


def main():
    parser = argparse.ArgumentParser(description="무신사 랭킹 크롤러")
    parser.add_argument("--category", default="000", help="카테고리 코드 (쉼표로 복수 지정 가능)")
    parser.add_argument("--period", default="REALTIME", choices=PERIODS.keys(), help="기간 필터")
    parser.add_argument("--gender", default="A", choices=GENDERS.keys(), help="성별 필터")
    parser.add_argument("--age", default="AGE_BAND_ALL", choices=AGE_BANDS.keys(), help="연령대 필터")
    parser.add_argument("--output-dir", default=None, help="저장 디렉토리")
    parser.add_argument("--format", default="xlsx", choices=["xlsx", "json", "both"], help="저장 형식")
    parser.add_argument("--download-images", action="store_true", help="상품 이미지 별도 다운로드")
    parser.add_argument("--no-images", action="store_true", help="엑셀 내 이미지 삽입 비활성화 (빠른 저장)")
    parser.add_argument("--list-categories", action="store_true", help="카테고리 목록 출력")
    args = parser.parse_args()

    if args.list_categories:
        list_categories()
        return

    # 저장 디렉토리 설정
    if args.output_dir:
        output_dir = args.output_dir
    else:
        # 프로젝트 루트/workspace 기준
        _d = os.path.dirname(os.path.abspath(__file__))
        for _ in range(5):
            if os.path.isfile(os.path.join(_d, "CLAUDE.md")):
                break
            _d = os.path.dirname(_d)
        project_root = _d
        today = datetime.now().strftime("%Y%m%d")
        output_dir = os.path.join(project_root, "workspace", "musinsa-ranking", f"ranking_{today}")

    os.makedirs(output_dir, exist_ok=True)

    # 복수 카테고리 처리
    categories = [c.strip() for c in args.category.split(",")]
    all_products = []

    print(f"\n{'='*60}")
    print(f"  무신사 랭킹 크롤러")
    print(f"  카테고리: {', '.join(CATEGORIES.get(c, c) for c in categories)}")
    print(f"  기간: {PERIODS[args.period]} | 성별: {GENDERS[args.gender]} | 연령: {AGE_BANDS[args.age]}")
    print(f"{'='*60}\n")

    for cat in categories:
        cat_name = CATEGORIES.get(cat, cat)
        print(f"\n[{cat_name} ({cat})] 수집 중...")
        products = fetch_ranking(cat, args.period, args.gender, args.age)
        for p in products:
            p["category_code"] = cat
            p["category_name"] = cat_name
        all_products.extend(products)

    if not all_products:
        print("\n수집된 상품이 없습니다.")
        return

    # 메타데이터
    meta = {
        "categories": {c: CATEGORIES.get(c, c) for c in categories},
        "period": args.period,
        "period_name": PERIODS[args.period],
        "gender": args.gender,
        "gender_name": GENDERS[args.gender],
        "age_band": args.age,
        "age_name": AGE_BANDS[args.age],
        "category_name": ", ".join(CATEGORIES.get(c, c) for c in categories),
        "total_products": len(all_products),
        "crawled_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
    }

    # 파일명 생성
    cat_str = "-".join(categories)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    base_name = f"musinsa_ranking_{cat_str}_{args.period}_{args.gender}_{timestamp}"

    # 저장
    embed = not args.no_images
    if args.format in ("xlsx", "both"):
        xlsx_path = os.path.join(output_dir, f"{base_name}.xlsx")
        save_to_excel(all_products, xlsx_path, meta, embed_images=embed)

    if args.format in ("json", "both"):
        json_path = os.path.join(output_dir, f"{base_name}.json")
        save_to_json(all_products, json_path, meta)

    # 이미지 다운로드 (선택)
    if args.download_images:
        print("\n이미지 다운로드 중...")
        download_images(all_products, output_dir)

    # 결과 요약 (JSON으로 출력 — 스킬에서 파싱용)
    result = {
        "status": "success",
        "total_products": len(all_products),
        "output_dir": output_dir,
        "files": [],
        "meta": meta,
    }
    if args.format in ("xlsx", "both"):
        result["files"].append(xlsx_path)
    if args.format in ("json", "both"):
        result["files"].append(json_path)

    print(f"\n{'='*60}")
    print(f"  수집 완료: 총 {len(all_products)}개 상품")
    print(f"  저장 위치: {output_dir}")
    print(f"{'='*60}")

    # 스킬 연동용 JSON 결과
    result_path = os.path.join(output_dir, "_result.json")
    with open(result_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)


if __name__ == "__main__":
    main()
