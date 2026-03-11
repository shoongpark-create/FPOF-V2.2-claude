#!/usr/bin/env python3
"""
FPOF Dashboard Data Extractor
엑셀 원본 데이터 → JSON → HTML 대시보드 자동 생성

사용법:
  python3 scripts/dashboard/extract_data.py
  python3 scripts/dashboard/extract_data.py --week W10
  python3 scripts/dashboard/extract_data.py --sales-dir output/dashboard/weekly\ sales --master-dir output/dashboard/product\ master
"""

import argparse
import json
import os
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

# ── 프로젝트 루트 ──
ROOT = Path(__file__).resolve().parents[2]
DASHBOARD_DIR = ROOT / "output" / "dashboard"
SALES_DIR = DASHBOARD_DIR / "weekly sales"
MASTER_DIR = DASHBOARD_DIR / "product master"
TEMPLATE_HTML = DASHBOARD_DIR / "wacky-willy-dashboard.html"
OUTPUT_HTML = DASHBOARD_DIR / "wacky-willy-dashboard.html"

BRAND_FILTER = "와키윌리"
MULTI_BRANDS = ["와키윌리", "커버낫", "리"]


def safe_float(v):
    """안전한 float 변환 — #VALUE!, None, str 모두 처리"""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0


def detect_weeks(sales_dir):
    """sales 디렉토리에서 사용 가능한 주차 파일 탐색"""
    weeks = []
    for f in sorted(Path(sales_dir).glob("Weekly_Sales_Review_W*.xlsx")):
        m = re.search(r"W(\d+)", f.name)
        if m:
            weeks.append(int(m.group(1)))
    return sorted(weeks)


def detect_latest_week(sales_dir):
    weeks = detect_weeks(sales_dir)
    return weeks[-1] if weeks else None


def parse_season_from_style(style_code):
    """WA2601xx→26SS, WA2504xx→25FW"""
    m = re.match(r"^(?:WA|MG)(\d{2})(\d{2})", style_code or "")
    if not m:
        return ""
    yr, ss = m.group(1), int(m.group(2))
    return f"{yr}{'SS' if ss <= 2 else 'FW'}"


# ═══════════════════════════════════════════
# PART 1: Product Master 추출
# ═══════════════════════════════════════════
def build_spec_lookup(master_dir):
    """모든 Product Master 파일에서 item_code → specific_1~5 사전 구축.
    의류 아이템만 해당 (용품 카테고리는 속성 없음)."""
    spec = {}
    for f in sorted(Path(master_dir).glob("Weekly_Product_Master_W*.xlsx")):
        if "~$" in f.name:
            continue
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb["weekly_TY"]
        headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        col = {h: i for i, h in enumerate(headers) if h}
        sp_cols = [col.get(f"specific_{j}") for j in range(1, 6)]
        ic_col = col.get("basic_item_code", 14)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 15:
                continue
            item = str(row[ic_col] or "")
            if not item.startswith("WA") or item in spec:
                continue
            vals = [row[c] if c is not None else None for c in sp_cols]
            has_any = any(v is not None and str(v).strip() and str(v).strip() != "0" for v in vals)
            if has_any:
                spec[item] = {
                    f"sp{j+1}": (str(vals[j]).strip() if vals[j] and str(vals[j]).strip() != "0" else "")
                    for j in range(5)
                }
        wb.close()
    print(f"  [Spec Lookup] {len(spec)}개 아이템 속성 사전 구축")
    return spec


def extract_product_master(master_file, master_dir=None):
    """Weekly_Product_Master_W*.xlsx → product items + filters + ins
    master_dir 지정 시 모든 주차 파일을 스캔하여 누락 시즌 보완 + 속성 사전 구축"""
    print(f"  [Product Master] {master_file}")

    # 속성 사전 (모든 주차 파일에서)
    spec_lookup = build_spec_lookup(master_dir) if master_dir else {}

    # 여러 파일에서 데이터 수집 (최신 파일 우선, 누락 시즌 보완)
    all_rows = []
    col = None
    master_files = sorted(Path(master_dir).glob("Weekly_Product_Master_W*.xlsx"), reverse=True) if master_dir else [master_file]
    seen_seasons = set()
    for mf in master_files:
        if "~$" in mf.name:
            continue
        wb = openpyxl.load_workbook(mf, read_only=True, data_only=True)
        ws = wb["weekly_TY"]
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(h).strip() if h else "" for h in rows[0]]
        if col is None:
            col = {h: i for i, h in enumerate(headers) if h}
        # 이 파일에 있는 시즌 확인
        file_seasons = set()
        wa_rows = []
        for row in rows[1:]:
            if not row or len(row) < 15:
                continue
            item = str(row[col.get("basic_item_code", 14)] or "")
            bs = str(row[col.get("basic_season", 12)] or "")
            if item.startswith("WA") and bs:
                file_seasons.add(bs)
                wa_rows.append(row)
        # 아직 수집 안 된 시즌의 행만 추가
        new_seasons = file_seasons - seen_seasons
        if new_seasons:
            for row in wa_rows:
                bs = str(row[col.get("basic_season", 12)] or "")
                if bs in new_seasons:
                    all_rows.append(row)
            seen_seasons |= new_seasons
            print(f"    ← {mf.name}: 시즌 {sorted(new_seasons)} 추가 ({len(wa_rows)}행 중 {sum(1 for r in wa_rows if str(r[col.get('basic_season',12)] or '') in new_seasons)}행)")
        wb.close()

    print(f"  총 {len(all_rows)}행, 시즌: {sorted(seen_seasons)}")

    # 필요한 컬럼 인덱스 찾기
    needed_fields = [
        "year_season", "gender", "type", "category_L", "category_M",
        "basic_item_code", "basic_season",
        "order_qty", "order_price_tag_amt",
        "period_in_qty", "period_in_price_tag_amt",
        "cum_in_qty", "cum_in_price_tag_amt",
        "cum_out_qty", "cum_out_price_tag_amt",
        "period_sales_qty", "period_sales_price_tag_amt", "period_sales_actual_amt",
        "period_sales_7d_qty", "period_sales_7d_price_tag_amt",
        "cum_sales_qty", "cum_sales_price_tag_amt", "cum_sales_actual_amt",
        "total_stock_qty", "total_stock_price_tag_amt",
        "profit_pre_amt", "rate_sales_in_qty", "rate_sales_in_amt",
        "rate_sales_out_qty",
    ]

    # 아이템별 집계 (SKU → item_code)
    items_agg = defaultdict(lambda: {
        "order_qty": 0, "order_amt": 0, "order_price_tag_amt": 0,
        "period_in_qty": 0, "period_in_amt": 0, "period_in_price_tag_amt": 0,
        "cum_in_qty": 0, "cum_in_amt": 0, "cum_in_price_tag_amt": 0,
        "cum_out_qty": 0, "cum_out_amt": 0, "cum_out_price_tag_amt": 0,
        "period_sales_qty": 0, "period_sales_amt": 0, "period_sales_price_tag_amt": 0,
        "period_sales_actual_amt": 0,
        "period_sales_7d_qty": 0, "period_sales_7d_amt": 0, "period_sales_7d_price_tag_amt": 0,
        "period_sales_7d_actual_amt": 0,
        "cum_sales_qty": 0, "cum_sales_amt": 0, "cum_sales_price_tag_amt": 0,
        "cum_sales_actual_amt": 0,
        "total_stock_qty": 0, "total_stock_amt": 0, "total_stock_price_tag_amt": 0,
        "profit_pre_amt": 0, "profit_post_amt": 0,
        "_sku": 0,
    })
    item_meta = {}

    for row in all_rows:
        if not row or len(row) < 15:
            continue
        # basic_season 으로 시즌 판별
        basic_season = str(row[col.get("basic_season", 12)] or "")
        if not basic_season:
            continue

        # 와키윌리 필터: item_code가 WA로 시작
        item_code = str(row[col.get("basic_item_code", 14)] or "")
        if not item_code.startswith("WA"):
            continue

        year_season = basic_season  # e.g. "26SS", "25FW"
        gender = str(row[col.get("gender", 2)] or "")
        sub_season = str(row[col.get("season", 1)] or "")  # 봄/여름/가을/겨울
        cat_l = str(row[col.get("category_L", 4)] or "")
        cat_m = str(row[col.get("category_M", 5)] or "")

        key = item_code
        d = items_agg[key]
        d["_sku"] += 1

        # 메타데이터 (첫 만남에만)
        if key not in item_meta:
            sp = spec_lookup.get(item_code, {})
            item_meta[key] = {
                "year_season": year_season,
                "gender": gender,
                "sub_season": sub_season,
                "type": str(row[col.get("type", 3)] or ""),
                "category_L": cat_l,
                "category_M": cat_m,
                "item_code": item_code,
                "sp1": sp.get("sp1", ""),
                "sp2": sp.get("sp2", ""),
                "sp3": sp.get("sp3", ""),
                "sp4": sp.get("sp4", ""),
                "sp5": sp.get("sp5", ""),
            }

        # 수량/금액 집계
        field_map = {
            "order_qty": "order_qty",
            "order_price_tag_amt": "order_price_tag_amt",
            "period_in_qty": "period_in_qty",
            "period_in_price_tag_amt": "period_in_price_tag_amt",
            "cum_in_qty": "cum_in_qty",
            "cum_in_price_tag_amt": "cum_in_price_tag_amt",
            "cum_out_qty": "cum_out_qty",
            "cum_out_price_tag_amt": "cum_out_price_tag_amt",
            "period_sales_qty": "period_sales_qty",
            "period_sales_price_tag_amt": "period_sales_price_tag_amt",
            "period_sales_actual_amt": "period_sales_actual_amt",
            "period_sales_7d_qty": "period_sales_7d_qty",
            "period_sales_7d_price_tag_amt": "period_sales_7d_price_tag_amt",
            "cum_sales_qty": "cum_sales_qty",
            "cum_sales_price_tag_amt": "cum_sales_price_tag_amt",
            "cum_sales_actual_amt": "cum_sales_actual_amt",
            "total_stock_qty": "total_stock_qty",
            "total_stock_price_tag_amt": "total_stock_price_tag_amt",
            "profit_pre_amt": "profit_pre_amt",
        }

        for src, dst in field_map.items():
            if src in col:
                d[dst] += safe_float(row[col[src]])

    # 최종 items 리스트 구성 (경량화: 정수 반올림 + 미사용 필드 제거)
    items = []
    drop_fields = {"order_amt", "period_in_amt", "cum_in_amt", "cum_out_amt",
                    "period_sales_amt", "period_sales_7d_amt", "cum_sales_amt",
                    "total_stock_amt", "profit_post_amt", "period_sales_7d_actual_amt",
                    "category_L"}
    for key, d in items_agg.items():
        meta = item_meta[key]
        item = {**d, **meta}
        # 비율 계산
        in_q = item["cum_in_qty"]
        s_q = item["cum_sales_qty"]
        out_q = item["cum_out_qty"]
        sa = item["cum_sales_actual_amt"]
        pa = item["profit_pre_amt"]
        item["ri_q"] = round((s_q / in_q * 100), 1) if in_q else 0
        item["ri_a"] = round((item["cum_sales_price_tag_amt"] / item["cum_in_price_tag_amt"] * 100), 1) if item["cum_in_price_tag_amt"] else 0
        item["ro_q"] = round((s_q / out_q * 100), 1) if out_q else 0
        item["pr"] = round((pa / sa * 100), 1) if sa else 0
        # 금액/수량 정수화
        for k in list(item.keys()):
            if k in drop_fields:
                del item[k]
            elif isinstance(item[k], float):
                if k in ("ri_q", "ri_a", "ro_q", "pr"):
                    pass  # 이미 round 처리
                else:
                    item[k] = int(item[k])
        # 빈 spec 필드 제거
        for sp in ["sp1", "sp2", "sp3", "sp4", "sp5"]:
            if not item.get(sp):
                item.pop(sp, None)
        items.append(item)

    # 필터 옵션
    filters = {
        "year_season": sorted(set(i["year_season"] for i in items)),
        "gender": sorted(set(i["gender"] for i in items)),
        "category_M": sorted(set(i["category_M"] for i in items)),
        "sub_season": sorted(set(i.get("sub_season", "") for i in items if i.get("sub_season"))),
    }

    # ins (시즌별 집계)
    ins = {}
    for ss in filters["year_season"]:
        ss_items = [i for i in items if i["year_season"] == ss]
        if not ss_items:
            continue
        total = len(ss_items)
        sorted_by_sales = sorted(ss_items, key=lambda x: x["cum_sales_actual_amt"], reverse=True)
        total_sales = sum(i["cum_sales_actual_amt"] for i in ss_items)
        t20n = max(1, int(total * 0.2))
        t20_sales = sum(i["cum_sales_actual_amt"] for i in sorted_by_sales[:t20n])
        hit_rate = (t20_sales / total_sales * 100) if total_sales else 0

        total_in = sum(i["cum_in_qty"] for i in ss_items)
        total_sold = sum(i["cum_sales_qty"] for i in ss_items)
        acc_r = (total_sold / total_in * 100) if total_in else 0

        wms_items = [i for i in ss_items if i["gender"] == "우먼스"]
        wms_in = sum(i["cum_in_qty"] for i in wms_items)
        wms_sold = sum(i["cum_sales_qty"] for i in wms_items)
        wms_acc_r = (wms_sold / wms_in * 100) if wms_in else 0

        # 카테고리별 집계 (cp)
        cp = defaultdict(lambda: {"sa": 0, "sq": 0, "n": 0})
        for i in ss_items:
            c = i["category_M"]
            cp[c]["sa"] += i["cum_sales_actual_amt"]
            cp[c]["sq"] += i["cum_sales_qty"]
            cp[c]["n"] += 1

        # 히어로 후보 (카테고리별 top 아이템)
        hero = defaultdict(list)
        for i in sorted_by_sales:
            ri = i["ri_q"]
            if ri >= 50 and i["cum_sales_actual_amt"] > 0:
                hero[i["category_M"]].append({
                    "ic": i["item_code"],
                    "g": i["gender"],
                    "sa": i["cum_sales_actual_amt"],
                    "sq": i["cum_sales_qty"],
                    "r": ri,
                })

        ins[ss] = {
            "hit": round(hit_rate, 1),
            "t20n": t20n,
            "total": total,
            "acc_r": round(acc_r, 1),
            "wms_acc_r": round(wms_acc_r, 1),
            "cp": dict(cp),
            "hero": dict(hero),
        }

    return {"filters": filters, "ins": ins, "items": items}


# ═══════════════════════════════════════════
# PART 2: Weekly Sales 추출
# ═══════════════════════════════════════════
def extract_sales_sheet(wb, sheet_name, brand_filter=BRAND_FILTER):
    """단일 시트에서 브랜드 데이터 추출. brand_filter=None이면 전체."""
    if sheet_name not in wb.sheetnames:
        return []

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = rows[0]
    col = {str(h).strip(): i for i, h in enumerate(headers) if h}

    data = []
    for row in rows[1:]:
        if not row or len(row) < 10:
            continue
        brand = str(row[col.get("Brand", 1)] or "")
        if brand_filter is not None and brand != brand_filter:
            continue

        d = {
            "brand": brand,
            "month": str(row[col.get("Sales Month", 0)] or ""),
            "item": str(row[col.get("Item", 2)] or ""),
            "gender": str(row[col.get("Uni/Wms/Kids", 3)] or ""),
            "year_season": str(row[col.get("Year+Season", 7)] or ""),
            "channel": str(row[col.get("Channel", 9)] or ""),
            "channel_type": str(row[col.get("Channel type", col.get("Channel type", 20))] or ""),
            "store_code": str(row[col.get("Store", 10)] or ""),
            "store_name": str(row[col.get("Store Name", 11)] or ""),
            "style": str(row[col.get("Style No.", 12)] or ""),
            "style_name": str(row[col.get("Style Name", 13)] or ""),
            "category": str(row[col.get("Category", col.get("Category", 21))] or ""),
            "qty": safe_float(row[col.get("Sales Qty", 16)]),
            "net": safe_float(row[col.get("Sales Amt_Net", 17)]),
            "tag": safe_float(row[col.get(" Sales  Amt_TAG", col.get("Sales Amt_TAG", 18))]),
            "cost": safe_float(row[col.get("Cost Amt(V+)", 19)]),
        }
        data.append(d)
    return data


def agg_by_key(data, key_fn, filter_fn=None):
    """데이터를 key 기준으로 집계 (정수 반올림)"""
    result = defaultdict(lambda: {"qty": 0, "net": 0, "tag": 0, "cost": 0})
    for d in data:
        if filter_fn and not filter_fn(d):
            continue
        k = key_fn(d)
        if not k:
            continue
        result[k]["qty"] += d["qty"]
        result[k]["net"] += d["net"]
        result[k]["tag"] += d["tag"]
        result[k]["cost"] += d["cost"]
    # 정수 반올림 (JSON 크기 절감)
    for v in result.values():
        v["qty"] = int(v["qty"])
        v["net"] = int(v["net"])
        v["tag"] = int(v["tag"])
        v["cost"] = int(v["cost"])
    return dict(result)


def extract_sales_data(sales_dir, weeks, latest_week):
    """모든 주차의 Sales 데이터를 추출하여 통합"""
    print(f"  [Sales] 주차: {weeks}, 최신: W{latest_week}")

    latest_file = Path(sales_dir) / f"Weekly_Sales_Review_W{latest_week}.xlsx"
    print(f"  [Sales] Loading {latest_file.name}...")
    wb = openpyxl.load_workbook(latest_file, read_only=True, data_only=True)

    # ── cum_TY / cum_LY ──
    cum_ty = extract_sales_sheet(wb, "cum_TY")
    cum_ly = extract_sales_sheet(wb, "cum_LY")

    # ── period_TY / period_LY ──
    per_ty = extract_sales_sheet(wb, "period_TY")
    per_ly = extract_sales_sheet(wb, "period_LY")
    wb.close()

    # ── WoW: 이전 주차의 period 데이터 ──
    wow_data = []
    prev_week = latest_week - 1
    if prev_week in weeks:
        prev_file = Path(sales_dir) / f"Weekly_Sales_Review_W{prev_week}.xlsx"
        print(f"  [Sales] Loading W{prev_week} for WoW...")
        wb2 = openpyxl.load_workbook(prev_file, read_only=True, data_only=True)
        wow_data = extract_sales_sheet(wb2, "period_TY")
        wb2.close()

    # ── 집계 ──
    sales = {
        "cum": {
            "ch_ty": agg_by_key(cum_ty, lambda d: d["channel"]),
            "ch_ly": agg_by_key(cum_ly, lambda d: d["channel"]),
            "ct_ty": agg_by_key(cum_ty, lambda d: d["channel_type"]),
            "ct_ly": agg_by_key(cum_ly, lambda d: d["channel_type"]),
            "mo_ty": agg_by_key(cum_ty, lambda d: d["month"]),
            "mo_ly": agg_by_key(cum_ly, lambda d: d["month"]),
            "gd_ty": agg_by_key(cum_ty, lambda d: d["gender"]),
            "gd_ly": agg_by_key(cum_ly, lambda d: d["gender"]),
            "cat_ty": agg_by_key(cum_ty, lambda d: d["category"]),
            "cat_ly": agg_by_key(cum_ly, lambda d: d["category"]),
            "item_ty": agg_by_key(cum_ty, lambda d: d["item"]),
            "item_ly": agg_by_key(cum_ly, lambda d: d["item"]),
        },
        "per": {
            "ch_ty": agg_by_key(per_ty, lambda d: d["channel"]),
            "ch_ly": agg_by_key(per_ly, lambda d: d["channel"]),
            "ct_ty": agg_by_key(per_ty, lambda d: d["channel_type"]),
            "ct_ly": agg_by_key(per_ly, lambda d: d["channel_type"]),
            "gd_ty": agg_by_key(per_ty, lambda d: d["gender"]),
            "gd_ly": agg_by_key(per_ly, lambda d: d["gender"]),
            "cat_ty": agg_by_key(per_ty, lambda d: d["category"]),
            "cat_ly": agg_by_key(per_ly, lambda d: d["category"]),
            "item_ty": agg_by_key(per_ty, lambda d: d["item"]),
            "item_ly": agg_by_key(per_ly, lambda d: d["item"]),
        },
        "wow": {
            "ch_ty": agg_by_key(wow_data, lambda d: d["channel"]),
            "ct_ty": agg_by_key(wow_data, lambda d: d["channel_type"]),
            "gd_ty": agg_by_key(wow_data, lambda d: d["gender"]),
            "cat_ty": agg_by_key(wow_data, lambda d: d["category"]),
        },
    }

    # ── 매장 Top 20 (누적/주간) + 매장별 스타일 Top 20 ──
    store_agg = defaultdict(lambda: {"s": "", "n": "", "c": "", "net": 0, "qty": 0})
    store_style_agg = defaultdict(lambda: defaultdict(lambda: {"st": "", "nm": "", "it": "", "net": 0, "qty": 0}))
    for d in cum_ty:
        k = d["store_code"]
        store_agg[k]["s"] = k
        store_agg[k]["n"] = d["store_name"]
        store_agg[k]["c"] = d["channel"]
        store_agg[k]["net"] += d["net"]
        store_agg[k]["qty"] += d["qty"]
        # per-store style
        sk = d["style"]
        store_style_agg[k][sk]["st"] = sk
        store_style_agg[k][sk]["nm"] = d["style_name"]
        store_style_agg[k][sk]["it"] = d["item"]
        store_style_agg[k][sk]["net"] += d["net"]
        store_style_agg[k][sk]["qty"] += d["qty"]
    # 정수 반올림
    for v in store_agg.values():
        v["net"] = int(v["net"]); v["qty"] = int(v["qty"])
    stores = sorted(store_agg.values(), key=lambda x: x["net"], reverse=True)[:30]
    store_styles_cum = {}
    for st in stores:
        sc = st["s"]
        for sv in store_style_agg[sc].values():
            sv["net"] = int(sv["net"]); sv["qty"] = int(sv["qty"])
        top = sorted(store_style_agg[sc].values(), key=lambda x: x["net"], reverse=True)[:15]
        store_styles_cum[sc] = top

    pstore_agg = defaultdict(lambda: {"s": "", "n": "", "c": "", "net": 0, "qty": 0})
    pstore_style_agg = defaultdict(lambda: defaultdict(lambda: {"st": "", "nm": "", "it": "", "net": 0, "qty": 0}))
    for d in per_ty:
        k = d["store_code"]
        pstore_agg[k]["s"] = k
        pstore_agg[k]["n"] = d["store_name"]
        pstore_agg[k]["c"] = d["channel"]
        pstore_agg[k]["net"] += d["net"]
        pstore_agg[k]["qty"] += d["qty"]
        sk = d["style"]
        pstore_style_agg[k][sk]["st"] = sk
        pstore_style_agg[k][sk]["nm"] = d["style_name"]
        pstore_style_agg[k][sk]["it"] = d["item"]
        pstore_style_agg[k][sk]["net"] += d["net"]
        pstore_style_agg[k][sk]["qty"] += d["qty"]
    for v in pstore_agg.values():
        v["net"] = int(v["net"]); v["qty"] = int(v["qty"])
    per_stores = sorted(pstore_agg.values(), key=lambda x: x["net"], reverse=True)[:30]
    store_styles_per = {}
    for st in per_stores:
        sc = st["s"]
        for sv in pstore_style_agg[sc].values():
            sv["net"] = int(sv["net"]); sv["qty"] = int(sv["qty"])
        top = sorted(pstore_style_agg[sc].values(), key=lambda x: x["net"], reverse=True)[:15]
        store_styles_per[sc] = top

    sales["stores"] = stores
    sales["per_stores"] = per_stores
    sales["store_styles_cum"] = store_styles_cum
    sales["store_styles_per"] = store_styles_per

    # ── 스타일 Top (누적) ──
    style_agg = defaultdict(lambda: {"st": "", "nm": "", "it": "", "gd": "", "net": 0, "qty": 0, "tag": 0})
    for d in cum_ty:
        k = d["style"]
        style_agg[k]["st"] = k
        style_agg[k]["nm"] = d["style_name"]
        style_agg[k]["it"] = d["item"]
        style_agg[k]["gd"] = d["gender"]
        style_agg[k]["net"] += d["net"]
        style_agg[k]["qty"] += d["qty"]
        style_agg[k]["tag"] += d["tag"]
    for v in style_agg.values():
        v["net"] = int(v["net"]); v["qty"] = int(v["qty"]); v["tag"] = int(v["tag"])
    styles = sorted(style_agg.values(), key=lambda x: x["net"], reverse=True)[:50]
    sales["styles"] = styles

    # ── 스타일 Top (주간) ──
    pstyle_agg = defaultdict(lambda: {"st": "", "nm": "", "it": "", "gd": "", "net": 0, "qty": 0, "tag": 0})
    for d in per_ty:
        k = d["style"]
        pstyle_agg[k]["st"] = k
        pstyle_agg[k]["nm"] = d["style_name"]
        pstyle_agg[k]["it"] = d["item"]
        pstyle_agg[k]["gd"] = d["gender"]
        pstyle_agg[k]["net"] += d["net"]
        pstyle_agg[k]["qty"] += d["qty"]
        pstyle_agg[k]["tag"] += d["tag"]
    for v in pstyle_agg.values():
        v["net"] = int(v["net"]); v["qty"] = int(v["qty"]); v["tag"] = int(v["tag"])
    per_styles = sorted(pstyle_agg.values(), key=lambda x: x["net"], reverse=True)[:50]

    # ── 주차별 합계 (wt) ──
    wt = {}
    for w in weeks:
        wf = Path(sales_dir) / f"Weekly_Sales_Review_W{w}.xlsx"
        if not wf.exists():
            continue
        print(f"  [Sales] Loading W{w} for weekly totals...")
        wb_w = openpyxl.load_workbook(wf, read_only=True, data_only=True)

        # Cum TY/LY
        w_cum_ty = extract_sales_sheet(wb_w, "cum_TY")
        w_cum_ly = extract_sales_sheet(wb_w, "cum_LY")
        w_per_ty = extract_sales_sheet(wb_w, "period_TY")
        w_per_ly = extract_sales_sheet(wb_w, "period_LY")
        wb_w.close()

        def total(data):
            return {
                "qty": int(sum(d["qty"] for d in data)),
                "net": int(sum(d["net"] for d in data)),
                "tag": int(sum(d["tag"] for d in data)),
                "cost": int(sum(d["cost"] for d in data)),
            }

        wt[f"W{w}_cT"] = total(w_cum_ty)
        wt[f"W{w}_cL"] = total(w_cum_ly)
        wt[f"W{w}_pT"] = total(w_per_ty)
        wt[f"W{w}_pL"] = total(w_per_ly)

        # 주차별 아이템(세부카테고리) 매출 (period TY/LY)
        wt[f"W{w}_item_pT"] = agg_by_key(w_per_ty, lambda d: d["item"])
        wt[f"W{w}_item_pL"] = agg_by_key(w_per_ly, lambda d: d["item"])

    sales["wt"] = wt

    # ── 스타일 상세 (채널/매장 드릴다운) ──
    # Top 100 스타일만 상세 데이터 생성 (크기 절감)
    top_styles_set = set(s["st"] for s in styles[:100])
    top_pstyles_set = set(s["st"] for s in per_styles[:100])
    all_detail_styles = top_styles_set | top_pstyles_set

    cum_detail = {}
    per_detail = {}

    cum_by_style = defaultdict(list)
    for d in cum_ty:
        if d["style"] in all_detail_styles:
            cum_by_style[d["style"]].append(d)

    for st, st_data in cum_by_style.items():
        ch_agg = defaultdict(lambda: {"c": "", "net": 0, "qty": 0})
        store_agg2 = defaultdict(lambda: {"s": "", "n": "", "c": "", "net": 0, "qty": 0})
        for d in st_data:
            ch_agg[d["channel"]]["c"] = d["channel"]
            ch_agg[d["channel"]]["net"] += d["net"]
            ch_agg[d["channel"]]["qty"] += d["qty"]
            store_agg2[d["store_code"]]["s"] = d["store_code"]
            store_agg2[d["store_code"]]["n"] = d["store_name"]
            store_agg2[d["store_code"]]["c"] = d["channel"]
            store_agg2[d["store_code"]]["net"] += d["net"]
            store_agg2[d["store_code"]]["qty"] += d["qty"]
        for v in ch_agg.values():
            v["net"] = int(v["net"]); v["qty"] = int(v["qty"])
        for v in store_agg2.values():
            v["net"] = int(v["net"]); v["qty"] = int(v["qty"])
        cum_detail[st] = {
            "ch": sorted(ch_agg.values(), key=lambda x: x["net"], reverse=True),
            "top": sorted(store_agg2.values(), key=lambda x: x["net"], reverse=True)[:8],
        }

    per_by_style = defaultdict(list)
    for d in per_ty:
        if d["style"] in all_detail_styles:
            per_by_style[d["style"]].append(d)

    for st, st_pdata in per_by_style.items():
        pch_agg = defaultdict(lambda: {"c": "", "net": 0, "qty": 0})
        pstore_agg2 = defaultdict(lambda: {"s": "", "n": "", "c": "", "net": 0, "qty": 0})
        for d in st_pdata:
            pch_agg[d["channel"]]["c"] = d["channel"]
            pch_agg[d["channel"]]["net"] += d["net"]
            pch_agg[d["channel"]]["qty"] += d["qty"]
            pstore_agg2[d["store_code"]]["s"] = d["store_code"]
            pstore_agg2[d["store_code"]]["n"] = d["store_name"]
            pstore_agg2[d["store_code"]]["c"] = d["channel"]
            pstore_agg2[d["store_code"]]["net"] += d["net"]
            pstore_agg2[d["store_code"]]["qty"] += d["qty"]
        for v in pch_agg.values():
            v["net"] = int(v["net"]); v["qty"] = int(v["qty"])
        for v in pstore_agg2.values():
            v["net"] = int(v["net"]); v["qty"] = int(v["qty"])
        per_detail[st] = {
            "ch": sorted(pch_agg.values(), key=lambda x: x["net"], reverse=True),
            "top": sorted(pstore_agg2.values(), key=lambda x: x["net"], reverse=True)[:8],
        }

    # ── 멀티 브랜드 채널 데이터 (와키윌리/커버낫/리) ──
    latest_file = Path(sales_dir) / f"Weekly_Sales_Review_W{latest_week}.xlsx"
    wb_mb = openpyxl.load_workbook(latest_file, read_only=True, data_only=True)
    all_cum_ty = extract_sales_sheet(wb_mb, "cum_TY", brand_filter=None)
    all_cum_ly = extract_sales_sheet(wb_mb, "cum_LY", brand_filter=None)
    all_per_ty = extract_sales_sheet(wb_mb, "period_TY", brand_filter=None)
    all_per_ly = extract_sales_sheet(wb_mb, "period_LY", brand_filter=None)
    wb_mb.close()

    # WoW 멀티브랜드
    all_wow_data = []
    if prev_week in weeks:
        prev_file = Path(sales_dir) / f"Weekly_Sales_Review_W{prev_week}.xlsx"
        wb_mb2 = openpyxl.load_workbook(prev_file, read_only=True, data_only=True)
        all_wow_data = extract_sales_sheet(wb_mb2, "period_TY", brand_filter=None)
        wb_mb2.close()

    multi_brand = {}
    for brand in MULTI_BRANDS:
        b_cum_ty = [d for d in all_cum_ty if d["brand"] == brand]
        b_cum_ly = [d for d in all_cum_ly if d["brand"] == brand]
        b_per_ty = [d for d in all_per_ty if d["brand"] == brand]
        b_per_ly = [d for d in all_per_ly if d["brand"] == brand]
        b_wow = [d for d in all_wow_data if d["brand"] == brand]

        # 매출 합계
        b_cum_net = sum(d["net"] for d in b_cum_ty)
        b_cum_net_ly = sum(d["net"] for d in b_cum_ly)
        b_per_net = sum(d["net"] for d in b_per_ty)
        b_per_net_ly = sum(d["net"] for d in b_per_ly)
        b_wow_net = sum(d["net"] for d in b_wow)
        b_cum_tag = sum(d["tag"] for d in b_cum_ty)
        b_cum_tag_ly = sum(d["tag"] for d in b_cum_ly)

        # 채널 집계
        b_ch = {
            "cum": {
                "ch_ty": agg_by_key(b_cum_ty, lambda d: d["channel"]),
                "ch_ly": agg_by_key(b_cum_ly, lambda d: d["channel"]),
                "ct_ty": agg_by_key(b_cum_ty, lambda d: d["channel_type"]),
                "ct_ly": agg_by_key(b_cum_ly, lambda d: d["channel_type"]),
            },
            "per": {
                "ch_ty": agg_by_key(b_per_ty, lambda d: d["channel"]),
                "ch_ly": agg_by_key(b_per_ly, lambda d: d["channel"]),
                "ct_ty": agg_by_key(b_per_ty, lambda d: d["channel_type"]),
                "ct_ly": agg_by_key(b_per_ly, lambda d: d["channel_type"]),
            },
            "wow": {
                "ch_ty": agg_by_key(b_wow, lambda d: d["channel"]),
            },
            "totals": {
                "cum_net": b_cum_net, "cum_net_ly": b_cum_net_ly,
                "per_net": b_per_net, "per_net_ly": b_per_net_ly,
                "wow_net": b_wow_net,
                "cum_tag": b_cum_tag, "cum_tag_ly": b_cum_tag_ly,
            },
        }

        # 매장 Top 20 (누적)
        b_store_agg = defaultdict(lambda: {"s": "", "n": "", "c": "", "net": 0, "qty": 0})
        b_store_style = defaultdict(lambda: defaultdict(lambda: {"st": "", "nm": "", "it": "", "net": 0, "qty": 0}))
        for d in b_cum_ty:
            k = d["store_code"]
            b_store_agg[k]["s"] = k
            b_store_agg[k]["n"] = d["store_name"]
            b_store_agg[k]["c"] = d["channel"]
            b_store_agg[k]["net"] += d["net"]
            b_store_agg[k]["qty"] += d["qty"]
            sk = d["style"]
            b_store_style[k][sk]["st"] = sk
            b_store_style[k][sk]["nm"] = d["style_name"]
            b_store_style[k][sk]["it"] = d["item"]
            b_store_style[k][sk]["net"] += d["net"]
            b_store_style[k][sk]["qty"] += d["qty"]
        for v in b_store_agg.values():
            v["net"] = int(v["net"]); v["qty"] = int(v["qty"])
        b_stores = sorted(b_store_agg.values(), key=lambda x: x["net"], reverse=True)[:20]
        b_ssc = {}
        for st in b_stores:
            sc = st["s"]
            for sv in b_store_style[sc].values():
                sv["net"] = int(sv["net"]); sv["qty"] = int(sv["qty"])
            b_ssc[sc] = sorted(b_store_style[sc].values(), key=lambda x: x["net"], reverse=True)[:10]
        b_ch["stores"] = b_stores
        b_ch["store_styles_cum"] = b_ssc

        # 매장 Top 20 (주간)
        b_pstore_agg = defaultdict(lambda: {"s": "", "n": "", "c": "", "net": 0, "qty": 0})
        b_pstore_style = defaultdict(lambda: defaultdict(lambda: {"st": "", "nm": "", "it": "", "net": 0, "qty": 0}))
        for d in b_per_ty:
            k = d["store_code"]
            b_pstore_agg[k]["s"] = k
            b_pstore_agg[k]["n"] = d["store_name"]
            b_pstore_agg[k]["c"] = d["channel"]
            b_pstore_agg[k]["net"] += d["net"]
            b_pstore_agg[k]["qty"] += d["qty"]
            sk = d["style"]
            b_pstore_style[k][sk]["st"] = sk
            b_pstore_style[k][sk]["nm"] = d["style_name"]
            b_pstore_style[k][sk]["it"] = d["item"]
            b_pstore_style[k][sk]["net"] += d["net"]
            b_pstore_style[k][sk]["qty"] += d["qty"]
        for v in b_pstore_agg.values():
            v["net"] = int(v["net"]); v["qty"] = int(v["qty"])
        b_pstores = sorted(b_pstore_agg.values(), key=lambda x: x["net"], reverse=True)[:20]
        b_ssp = {}
        for st in b_pstores:
            sc = st["s"]
            for sv in b_pstore_style[sc].values():
                sv["net"] = int(sv["net"]); sv["qty"] = int(sv["qty"])
            b_ssp[sc] = sorted(b_pstore_style[sc].values(), key=lambda x: x["net"], reverse=True)[:10]
        b_ch["per_stores"] = b_pstores
        b_ch["store_styles_per"] = b_ssp

        # 주차별 매출 추이
        b_wt = {}
        for w in weeks:
            wf = Path(sales_dir) / f"Weekly_Sales_Review_W{w}.xlsx"
            if not wf.exists():
                continue
            wb_w = openpyxl.load_workbook(wf, read_only=True, data_only=True)
            w_cum = [d for d in extract_sales_sheet(wb_w, "cum_TY", brand_filter=None) if d["brand"] == brand]
            w_per = [d for d in extract_sales_sheet(wb_w, "period_TY", brand_filter=None) if d["brand"] == brand]
            w_per_ly = [d for d in extract_sales_sheet(wb_w, "period_LY", brand_filter=None) if d["brand"] == brand]
            wb_w.close()
            b_wt[f"W{w}_cT"] = sum(d["net"] for d in w_cum)
            b_wt[f"W{w}_pT"] = sum(d["net"] for d in w_per)
            b_wt[f"W{w}_pL"] = sum(d["net"] for d in w_per_ly)
        b_ch["wt"] = b_wt

        multi_brand[brand] = b_ch

    sales["multi_brand"] = multi_brand

    return sales, per_styles, cum_detail, per_detail


# ═══════════════════════════════════════════
# PART 3: HTML 대시보드 생성
# ═══════════════════════════════════════════
def build_dashboard(prod_data, sales_data, per_styles, cum_detail, per_detail, latest_week, template_html):
    """기존 HTML 템플릿의 데이터 블록만 교체하여 업데이트

    전략: 라인 기반으로 데이터 영역을 찾아 교체
    - const __PROD_DATA__ = ... 부터
    - const __PER_DETAIL__ = ...; 까지를 새 데이터로 교체
    """
    print(f"  [HTML] Building dashboard...")

    lines = template_html.read_text(encoding="utf-8").split("\n")

    # 데이터 블록 시작/끝 라인 찾기
    data_start = None  # const __PROD_DATA__ 시작 라인
    data_end = None    # 마지막 데이터 변수 다음 라인 (로직 코드 시작)

    # 데이터 변수 목록 (순서대로)
    data_vars = ["__PROD_DATA__", "__SALES_DATA__", "__PER_STYLES__", "__CUM_DETAIL__", "__PER_DETAIL__"]

    for i, line in enumerate(lines):
        stripped = line.strip()
        if stripped.startswith("const __PROD_DATA__") and data_start is None:
            data_start = i
        # 마지막 데이터 변수 이후 첫 번째 비데이터 라인
        if data_start is not None and data_end is None:
            if not any(stripped.startswith(f"const {v}") for v in data_vars):
                # 빈줄이 아니고, 데이터 시작 이후이며, const __로 시작하지 않으면
                # 아직 데이터 변수의 값 부분일 수 있으므로 확인 필요
                pass

    # 더 정확한 방법: 각 데이터 변수의 시작 라인을 찾고, 마지막 변수 이후 로직 코드 시작점을 찾음
    var_lines = {}
    for i, line in enumerate(lines):
        for v in data_vars:
            if line.strip().startswith(f"const {v}"):
                var_lines[v] = i

    if not var_lines:
        print("  ERROR: 데이터 변수를 찾을 수 없습니다.")
        return "\n".join(lines)

    data_start = min(var_lines.values())
    last_var_line = max(var_lines.values())

    # 마지막 데이터 변수 이후, 다음 코드 라인 찾기
    # (데이터가 한 줄에 있을 수도, 여러 줄에 걸쳐있을 수도 있음)
    # "const CT=" 또는 "const COLORS=" 등 로직 코드의 시작을 찾음
    data_end = last_var_line + 1
    for i in range(last_var_line + 1, len(lines)):
        stripped = lines[i].strip()
        if not stripped:
            continue
        # 데이터 변수가 아닌 첫 const/function/var/let 등
        if any(stripped.startswith(p) for p in ["const ", "var ", "let ", "function ", "//"]):
            if not any(stripped.startswith(f"const {v}") for v in data_vars):
                data_end = i
                break

    print(f"  데이터 영역: L{data_start+1} ~ L{data_end} (교체 대상)")

    # 새 데이터 블록 생성
    new_data_lines = []
    new_data_lines.append(f"const __PROD_DATA__ = {json.dumps(prod_data, ensure_ascii=False, separators=(',', ':'))};")
    new_data_lines.append(f"const __SALES_DATA__ = {json.dumps(sales_data, ensure_ascii=False, separators=(',', ':'))};")
    new_data_lines.append(f"const __PER_STYLES__ = {json.dumps(per_styles, ensure_ascii=False, separators=(',', ':'))};")
    new_data_lines.append(f"const __CUM_DETAIL__ = {json.dumps(cum_detail, ensure_ascii=False, separators=(',', ':'))};")
    new_data_lines.append(f"const __PER_DETAIL__ = {json.dumps(per_detail, ensure_ascii=False, separators=(',', ':'))};")

    # 조립: 헤더(CSS+HTML) + 새 데이터 + 로직 코드
    result_lines = lines[:data_start] + new_data_lines + lines[data_end:]

    html = "\n".join(result_lines)

    # 헤더의 주차 정보 업데이트
    html = re.sub(r"W\d+ 기준", f"W{latest_week} 기준", html)

    return html


# ═══════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════
def main():
    parser = argparse.ArgumentParser(description="FPOF Dashboard Data Extractor")
    parser.add_argument("--week", type=str, help="처리할 최신 주차 (e.g. W10). 생략 시 자동 감지")
    parser.add_argument("--sales-dir", type=str, default=str(SALES_DIR))
    parser.add_argument("--master-dir", type=str, default=str(MASTER_DIR))
    parser.add_argument("--output", type=str, default=str(OUTPUT_HTML))
    parser.add_argument("--json-only", action="store_true", help="JSON 파일만 생성 (HTML 빌드 스킵)")
    args = parser.parse_args()

    sales_dir = Path(args.sales_dir)
    master_dir = Path(args.master_dir)

    # 주차 감지
    weeks = detect_weeks(sales_dir)
    if not weeks:
        print("ERROR: Weekly_Sales_Review_W*.xlsx 파일을 찾을 수 없습니다.")
        sys.exit(1)

    if args.week:
        latest_week = int(args.week.replace("W", "").replace("w", ""))
    else:
        latest_week = weeks[-1]

    print(f"═══ FPOF Dashboard Extractor ═══")
    print(f"  주차: {weeks} → 최신 W{latest_week}")
    print()

    # 1. Product Master 추출
    master_file = master_dir / f"Weekly_Product_Master_W{latest_week}.xlsx"
    if not master_file.exists():
        print(f"WARNING: {master_file} not found, trying latest available...")
        available_masters = sorted(master_dir.glob("Weekly_Product_Master_W*.xlsx"))
        if available_masters:
            master_file = available_masters[-1]
        else:
            print("ERROR: Product Master 파일을 찾을 수 없습니다.")
            sys.exit(1)

    prod_data = extract_product_master(master_file, master_dir=master_dir)
    print(f"  → {len(prod_data['items'])} items, {len(prod_data['ins'])} seasons")

    # 2. Sales 추출
    sales_data, per_styles, cum_detail, per_detail = extract_sales_data(sales_dir, weeks, latest_week)
    print(f"  → {len(sales_data['wt'])} weekly totals, {len(sales_data['styles'])} styles")
    print(f"  → {len(cum_detail)} cum details, {len(per_detail)} per details")

    # 3. JSON 저장
    out_dir = DASHBOARD_DIR
    json.dump(prod_data, open(out_dir / "product_status_data.json", "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    json.dump(sales_data, open(out_dir / "sales_data.json", "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    json.dump(per_styles, open(out_dir / "_per_styles_v2.json", "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    json.dump(cum_detail, open(out_dir / "_cum_detail_v2.json", "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    json.dump(per_detail, open(out_dir / "_per_detail_v2.json", "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    print(f"\n  ✓ JSON 파일 저장 완료 → {out_dir}")

    if args.json_only:
        print("  (--json-only: HTML 빌드 스킵)")
        return

    # 4. HTML 빌드
    if not TEMPLATE_HTML.exists():
        print(f"ERROR: 템플릿 HTML을 찾을 수 없습니다: {TEMPLATE_HTML}")
        sys.exit(1)

    html = build_dashboard(prod_data, sales_data, per_styles, cum_detail, per_detail, latest_week, TEMPLATE_HTML)
    output_path = Path(args.output)
    output_path.write_text(html, encoding="utf-8")
    print(f"  ✓ HTML 대시보드 생성 완료 → {output_path}")
    print(f"\n═══ 완료! W{latest_week} 기준 대시보드가 업데이트되었습니다 ═══")


if __name__ == "__main__":
    main()
